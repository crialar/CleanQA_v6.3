import streamlit as st
import zipfile
import io
import re
import difflib
import time
import html
import hashlib
import bisect
import numpy as np
from collections import Counter
from datetime import date
from typing import List, Dict
from anonymizer import MQXLIFFAnonymizer, load_dictionary_terms
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import qa_checker

st.set_page_config(
    page_title="Clean&QA",
    page_icon="🔒",
    layout="wide",
    initial_sidebar_state="expanded"
)

CUSTOM_CSS = """
<style>
    .main {
        background-color: #ffffff !important;
    }
    .stApp {
        background-color: #ffffff !important;
    }
    .stApp, .stApp p, .stApp span, .stApp label, .stApp div {
        color: #130e45 !important;
    }
    
    /* Global scrollbar styles */
    *::-webkit-scrollbar {
        width: 16px !important;
        height: 16px !important;
    }
    *::-webkit-scrollbar-track {
        background: #d0d3d4 !important;
        border-radius: 8px !important;
    }
    *::-webkit-scrollbar-thumb {
        background: linear-gradient(180deg, #1a5488, #0e7bc0) !important;
        border-radius: 8px !important;
        border: 2px solid #d0d3d4 !important;
    }
    *::-webkit-scrollbar-thumb:hover {
        background: linear-gradient(180deg, #134277, #1a5488) !important;
    }
    
    /* Firefox scrollbar */
    * {
        scrollbar-width: auto;
        scrollbar-color: #1a5488 #d0d3d4;
    }
    [data-testid="stSidebar"] {
        background-color: #e0e3e4 !important;
    }
    [data-testid="stSidebar"] p, [data-testid="stSidebar"] span, [data-testid="stSidebar"] label {
        color: #130e45 !important;
    }
    h1, h2, h3 {
        color: #1a5488 !important;
    }
    .stButton > button,
    [data-testid="stBaseButton-primary"] {
        background-color: #0e7bc0 !important;
        color: white !important;
        border: none;
        border-radius: 8px;
        padding: 0.5rem 1rem;
        font-weight: 800;
    }
    .stButton > button p,
    .stButton > button span,
    .stButton > button div,
    [data-testid="stBaseButton-primary"] p,
    [data-testid="stBaseButton-primary"] span {
        color: white !important;
        background-color: transparent !important;
        padding: 0 !important;
    }
    [data-testid="stBaseButton-primary"] {
        font-weight: 900 !important;
        font-size: 1.1rem !important;
    }
    .stButton > button:hover,
    [data-testid="stBaseButton-primary"]:hover {
        background-color: #134277 !important;
    }
    .stButton > button:disabled,
    .stButton > button[disabled],
    .stDownloadButton > button:disabled,
    .stDownloadButton > button[disabled],
    [data-testid="stBaseButton-primary"]:disabled,
    [data-testid="stBaseButton-secondary"]:disabled {
        background-color: #d9dcdd !important;
        color: #8a8d92 !important;
        cursor: not-allowed !important;
        opacity: 0.7 !important;
    }
    .stButton > button:disabled p,
    .stButton > button:disabled span,
    .stButton > button:disabled div,
    .stDownloadButton > button:disabled p,
    .stDownloadButton > button:disabled span,
    .stDownloadButton > button:disabled div,
    [data-testid="stBaseButton-primary"]:disabled p,
    [data-testid="stBaseButton-primary"]:disabled span,
    [data-testid="stBaseButton-secondary"]:disabled p,
    [data-testid="stBaseButton-secondary"]:disabled span {
        color: #8a8d92 !important;
    }
    .stButton > button:disabled:hover,
    .stDownloadButton > button:disabled:hover,
    [data-testid="stBaseButton-primary"]:disabled:hover,
    [data-testid="stBaseButton-secondary"]:disabled:hover {
        background-color: #d9dcdd !important;
    }
    .stat-card {
        background-color: #e0e3e4;
        border-radius: 10px;
        padding: 1rem;
        text-align: center;
        border-left: 4px solid #1a5488;
    }
    .stat-card-safe-regex { border-left-color: #0e7bc0; }
    .stat-card-regex-ct { border-left-color: #6f42c1; }
    .stat-card-proper-names { border-left-color: #e83e8c; }
    .stat-card-dictionary { border-left-color: #17a2b8; }
    .stat-number {
        font-size: 2rem;
        font-weight: bold;
        color: #130e45;
    }
    .stat-label {
        color: #5e5f6b;
        font-size: 0.9rem;
    }
    .sidebar-divider {
        border: none;
        border-top: 1px solid #c0c3c4;
        margin: 0.8rem 0;
    }
    .app-footer {
        text-align: center;
        padding: 1.5rem 0 0.5rem 0;
        color: #8a8b96 !important;
        font-size: 0.8rem;
        border-top: 1px solid #e0e3e4;
        margin-top: 2rem;
    }
    .app-footer p, .app-footer span {
        color: #8a8b96 !important;
    }
    .preview-box {
        background-color: #f8f9fa;
        border: 1px solid #bcbdbe;
        border-radius: 8px;
        padding: 1rem;
        margin: 0.5rem 0;
    }
    .preview-container {
        max-height: 500px;
        overflow-y: auto;
        padding-right: 10px;
        margin: 1rem 0;
        border: 1px solid #e0e3e4;
        border-radius: 8px;
        background-color: #fafafa;
    }
    .preview-container::-webkit-scrollbar {
        width: 18px;
    }
    .preview-container::-webkit-scrollbar-track {
        background: #d0d3d4;
        border-radius: 9px;
        box-shadow: inset 0 0 3px rgba(0,0,0,0.2);
    }
    .preview-container::-webkit-scrollbar-thumb {
        background: linear-gradient(180deg, #1a5488, #0e7bc0);
        border-radius: 9px;
        border: 3px solid #d0d3d4;
        box-shadow: 0 2px 4px rgba(0,0,0,0.3);
    }
    .preview-container::-webkit-scrollbar-thumb:hover {
        background: linear-gradient(180deg, #134277, #1a5488);
    }
    div[data-testid="stExpander"] {
        max-height: 600px;
        overflow-y: auto;
    }
    div[data-testid="stExpander"]::-webkit-scrollbar {
        width: 20px;
    }
    div[data-testid="stExpander"]::-webkit-scrollbar-track {
        background: #c8cbcc;
        border-radius: 10px;
        box-shadow: inset 0 0 4px rgba(0,0,0,0.25);
    }
    div[data-testid="stExpander"]::-webkit-scrollbar-thumb {
        background: linear-gradient(180deg, #1a5488, #0e7bc0);
        border-radius: 10px;
        border: 3px solid #c8cbcc;
        box-shadow: 0 2px 5px rgba(0,0,0,0.35);
    }
    div[data-testid="stExpander"]::-webkit-scrollbar-thumb:hover {
        background: linear-gradient(180deg, #134277, #1a5488);
    }
    .exclude-badge {
        background-color: #dc3545;
        color: white;
        padding: 0.2rem 0.5rem;
        border-radius: 4px;
        font-size: 0.75rem;
        font-weight: bold;
        margin-left: 0.5rem;
    }
    .excluded-segment {
        border: 2px dashed #dc3545 !important;
        background-color: #fff0f0 !important;
    }
    .before-text {
        color: #dc3545;
        background-color: #ffe6e6;
        padding: 0.5rem;
        border-radius: 4px;
    }
    .after-text {
        color: #28a745;
        background-color: #e6ffe6;
        padding: 0.5rem;
        border-radius: 4px;
    }
    .section-header {
        background-color: #7cb4db;
        color: white;
        padding: 0.5rem 1rem;
        border-radius: 8px;
        margin: 1rem 0;
    }
    .info-box {
        background-color: #e0e3e4;
        border-left: 4px solid #0e7bc0;
        padding: 1rem;
        border-radius: 0 8px 8px 0;
        margin: 1rem 0;
    }
    div[data-testid="stFileUploader"] {
        background-color: #f8f9fa;
        border: 2px dashed #7cb4db;
        border-radius: 10px;
        padding: 1rem;
    }
</style>
"""

st.markdown(CUSTOM_CSS, unsafe_allow_html=True)


def count_words(text: str) -> int:
    """Count words in text, ignoring XML tags and whitespace."""
    if not text:
        return 0
    import re
    clean = re.sub(r'<[^>]+>', ' ', text)
    clean = re.sub(r'\s+', ' ', clean).strip()
    words = [w for w in clean.split() if len(w) > 0]
    return len(words)


def segment_word_count(preview: dict) -> int:
    """Returns the max word count between source and target."""
    source_words = count_words(preview.get('source_before', ''))
    target_words = count_words(preview.get('target_before', ''))
    return max(source_words, target_words)


def is_junk_segment(preview: dict, min_words_junk: int = 3) -> bool:
    """Detects junk/short original segments that pollute TM databases.
    Returns True if the segment should be excluded."""
    import re
    source = preview.get('source_before', '').strip()
    target = preview.get('target_before', '').strip()
    
    if not source and not target:
        return True
    
    for text in [source, target]:
        if not text:
            continue
        clean = re.sub(r'<[^>]+>', ' ', text)
        clean = re.sub(r'\s+', ' ', clean).strip()
        if not clean:
            continue
        if re.fullmatch(r'[\d\s\-.,;:!?¿¡()\[\]{}/\\|@#$%^&*+=<>~`"\'°ºª•–—…\u2022\u2013\u2014\u2026]+', clean):
            return True
        words = [w for w in clean.split() if len(w) > 0]
        if len(words) < min_words_junk:
            return True
    
    return False


def bilingual_previews(previews: dict) -> dict:
    """TM pipeline consumers (dedup, clean TMX, quality funnel) only work
    with bilingual files. Word (.docx) previews are monolingual (empty
    source) and would otherwise collapse into a single giant conflict
    group / enter the TMX as empty-source TUs."""
    return {
        fn: fp for fn, fp in (previews or {}).items()
        if not fn.lower().endswith(".docx")
    }


def render_stat_card(label: str, value: int, col, css_class: str = ""):
    with col:
        st.markdown(f"""
        <div class="stat-card {css_class}">
            <div class="stat-number">{value}</div>
            <div class="stat-label">{label}</div>
        </div>
        """, unsafe_allow_html=True)


def _normalize_lang_code(code: str) -> str:
    """Normalize language code to standard format (e.g., de-de -> de-DE, en -> en)."""
    parts = code.strip().split("-")
    if len(parts) == 2:
        return f"{parts[0].lower()}-{parts[1].upper()}"
    return parts[0].lower()


def strip_inline_tags(text: str) -> str:
    if not text:
        return text
    cleaned = re.sub(r'<[^>]+>', '', text)
    cleaned = re.sub(r'\{/?(\d+)?\}', '', cleaned)
    cleaned = re.sub(r'  +', ' ', cleaned)
    return cleaned.strip()


def _html_escape(text: str) -> str:
    return text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")


# Single source of truth for the word-level diff renderer lives in
# qa_checker so both the Anonymizer Duplicates tab and the QA tab's
# `inconsistent_translation` card paint the same way (Task #51).
from qa_checker import highlight_diff  # noqa: E402


def _qa_highlight_search(html: str, pattern) -> str:
    """Overlay the user's QA search query on top of already-rendered HTML.

    ``html`` is the output of :func:`_qa_highlight` (so it may already
    contain ``<span>`` wrappers for the QA finding). ``pattern`` is the
    compiled regex from the search bar (or ``None`` / ``"invalid"`` when
    no search is active). Matches inside text nodes are wrapped in a
    yellow ``<mark>`` so the reviewer sees exactly where their query
    landed without losing the QA finding's color.

    The implementation walks the HTML by splitting on tag boundaries
    (``<…>``) and only paints inside text segments — so we never inject
    ``<mark>`` inside an attribute value or break an existing tag.
    """
    if not pattern or pattern == "invalid" or not html:
        return html
    wrap = (
        '<mark style="background:#fff59d;color:inherit;'
        'padding:0 1px;border-radius:2px;">{m}</mark>'
    )
    # First split on tag boundaries so we never paint inside attributes.
    parts = re.split(r"(<[^>]+>)", html)
    for i, part in enumerate(parts):
        if not part or part.startswith("<"):
            continue
        # Then split each text node on HTML entities (`&amp;`, `&lt;`, …)
        # so the regex never lands inside an entity reference and breaks
        # it (e.g. searching for "amp" or "&" must not corrupt `&amp;`).
        sub = re.split(r"(&[A-Za-z]+;|&#\d+;|&#x[0-9A-Fa-f]+;)", part)
        for j, chunk in enumerate(sub):
            if not chunk or (chunk.startswith("&") and chunk.endswith(";")):
                continue
            sub[j] = pattern.sub(lambda m: wrap.format(m=m.group(0)), chunk)
        parts[i] = "".join(sub)
    return "".join(parts)


def _qa_highlight(text: str, span=None, category_id: str = None) -> str:
    """HTML-escape ``text`` and wrap any occurrence of ``span`` in a highlight.

    ``span`` may be:
      * ``None`` / empty   -> plain escaped text.
      * a single ``str``   -> block highlight using the category's group
        color (Task #36: 8 group colors, one per QA group).
      * a ``list[str]``    -> per-token block highlight + underline (used
        by spell-check). Same group color as the block highlight; the
        extra underline disambiguates per-token matches at a glance.

    ``category_id`` selects the highlight color via
    :func:`qa_checker.get_highlight_color`. When omitted, falls back to
    the Content red palette so legacy callers keep their old look.
    """
    escaped = _html_escape(text or "")
    if not span:
        return escaped
    bg, fg = qa_checker.get_highlight_color(category_id)
    block_tpl = (
        f'<span style="background:{bg};color:{fg};font-weight:600;'
        f'padding:0 2px;border-radius:2px;">{{w}}</span>'
    )
    wavy_tpl = (
        f'<span style="background:{bg};color:{fg};font-weight:600;'
        f'padding:0 2px;border-radius:2px;text-decoration:underline;">{{w}}</span>'
    )
    # Task #65 — double_spaces highlight: ASCII spaces collapse in HTML,
    # so the painted span was invisible. Render the matched payload with
    # each space swapped for a middle dot `·` so the reviewer sees
    # exactly where (and how many) extra spaces live in the target.
    def _format_match(matched: str) -> str:
        if category_id == "double_spaces":
            return matched.replace(" ", "·")
        return matched
    if isinstance(span, (list, tuple)):
        tokens = [s for s in dict.fromkeys(span) if s]  # ordered dedupe
        if not tokens:
            return escaped
        tokens.sort(key=len, reverse=True)
        # Per-category boundary so short tokens don't paint inside larger
        # ones — e.g. "2" inside "28005" (number_mismatch), "AB1" inside
        # "AB12" (alphanum_id_mismatch), "http://a.com" inside
        # "http://a.com/path" (urls). spellcheck keeps its letter-only
        # boundary so accents/digit-suffixed words behave as expected.
        # See qa_checker.get_highlight_boundary for the rule.
        left, right = qa_checker.get_highlight_boundary(category_id)
        pattern = re.compile(
            left + r"(?:" +
            "|".join(re.escape(_html_escape(t)) for t in tokens) +
            r")" + right,
            re.IGNORECASE | re.UNICODE,
        )
        return pattern.sub(lambda m: wavy_tpl.format(w=_format_match(m.group(0))), escaped)
    span_escaped = _html_escape(span)
    if not span_escaped:
        return escaped
    # Mark only the FIRST occurrence (count=1). A global sub would also paint
    # incidental repeats — e.g. span="t" highlighting every "t" in "título del
    # estudio", or span="2" highlighting the "2" of "28005" in "2 March, CP
    # 28005". Checks paint the offending span and it almost always coincides
    # with the first occurrence (leading letter, differing number, unique ID).
    if category_id in ("glossary_violation", "forbidden_terms"):
        left, right = qa_checker.get_highlight_boundary(category_id)
        pattern = re.compile(
            left + re.escape(span_escaped) + right,
            re.IGNORECASE | re.UNICODE,
        )
    else:
        pattern = re.compile(re.escape(span_escaped), re.IGNORECASE)
    # Some checks (e.g. final_punctuation_mismatch) point at the LAST
    # occurrence of the span — the trailing '.' / '!' / '?'. Painting
    # the first match in a string like "1.2.3.4." would highlight the
    # wrong character. See qa_checker.LAST_OCCURRENCE_CATEGORIES.
    if category_id in qa_checker.LAST_OCCURRENCE_CATEGORIES:
        hits = list(pattern.finditer(escaped))
        if not hits:
            return escaped
        m = hits[-1]
        return (escaped[:m.start()]
                + block_tpl.format(w=_format_match(m.group(0)))
                + escaped[m.end():])
    return pattern.sub(lambda m: block_tpl.format(w=_format_match(m.group(0))), escaped, count=1)


def _qa_legend_html() -> str:
    """Render the Task #36 highlight-color legend (one swatch per QA group).

    Used at the top of the QA results panel so users can see at a glance
    which color belongs to which QA group.
    """
    swatches = []
    for grp, (bg, fg) in qa_checker.GROUP_HIGHLIGHT_COLORS.items():
        swatches.append(
            f'<span style="display:inline-block;background:{bg};'
            f'color:{fg};font-weight:600;padding:2px 8px;border-radius:3px;'
            f'margin:2px 4px 2px 0;font-size:0.78rem;">{_html_escape(grp)}</span>'
        )
    return (
        '<div style="margin:6px 0 10px 0;padding:8px 12px;background:#f5f7f8;'
        'border:1px solid #e0e3e4;border-radius:6px;font-size:0.85rem;">'
        '<strong style="color:#1a5488;">Highlight legend:</strong> '
        + "".join(swatches) + '</div>'
    )


_QA_SEG_FILTER_RE = __import__("re").compile(r"^[\d,\-\s]+$")


def _parse_segment_filter(query: str):
    """Return a set of segment IDs if *query* is a numeric segment-id filter
    (e.g. "12", "12,15,20", "10-20", "5, 8-11, 30"), else ``None`` so the
    caller falls back to a free-text search.

    Accepts digits, commas, hyphens (ranges) and whitespace; everything else
    means "treat as text search". Invalid ranges (start > end, empty parts)
    are silently dropped.
    """
    if not query:
        return None
    q = query.strip()
    if not q or not _QA_SEG_FILTER_RE.match(q):
        return None
    ids = set()
    for chunk in q.split(","):
        chunk = chunk.strip()
        if not chunk:
            continue
        if "-" in chunk:
            parts = [p.strip() for p in chunk.split("-") if p.strip()]
            if len(parts) != 2 or not all(p.isdigit() for p in parts):
                continue
            a, b = int(parts[0]), int(parts[1])
            if a > b:
                a, b = b, a
            ids.update(range(a, b + 1))
        elif chunk.isdigit():
            ids.add(int(chunk))
    return ids or None


def _qa_total_badge(count: int) -> str:
    """Render a neutral total-issue badge for the QA summary line.

    Task #68 dropped the HIGH/LOW severity surface: the per-card and
    summary-line pills are replaced by a single total. Brand-blue pill so
    it visually matches the rest of the QA toolbar without re-introducing
    a severity color.
    """
    return (f'<span style="background:#cce0ff;color:#1a5488;'
            f'border:1px solid #7cb4db;'
            f'padding:2px 10px;border-radius:12px;font-size:0.78rem;font-weight:700;'
            f'letter-spacing:0.02em;">TOTAL · {count}</span>')


def _qa_reset_segment_edit(seg_id):
    """Drop every per-card edit widget for one segment + its override entry.

    Used as an `on_click` callback so the widget keys are removed BEFORE
    the next script run instantiates the corresponding `st.text_area`s. On
    the next render every editor for this segment falls back to its
    original `value=` (the segment's untouched target text).

    Per-card widget keys follow the pattern ``qa_edit_{seg_id}__{card_uid}``
    so a single segment may have multiple sibling keys when the same
    segment surfaces in several QA categories.
    """
    prefix = f"qa_edit_{seg_id}__"
    for k in [k for k in list(st.session_state.keys()) if k.startswith(prefix)]:
        st.session_state.pop(k, None)
    overrides = st.session_state.get("qa_target_overrides", {})
    overrides.pop(str(seg_id), None)
    st.session_state["qa_target_overrides"] = overrides


def _qa_reset_all_edits():
    """Clear every inline-edit widget value and the entire override map."""
    for k in [k for k in list(st.session_state.keys()) if k.startswith("qa_edit_")]:
        st.session_state.pop(k, None)
    st.session_state["qa_target_overrides"] = {}



def _render_qa_check_tab():
    st.markdown("### 🛡️ QA Check")
    st.markdown(
        "<small>Standalone QA check for TMX/MQXLIFF files (or Word .docx, monolingual). No anonymization performed.</small>",
        unsafe_allow_html=True,
    )

    # ------------------------------------------------------------------
    # Shortcut: load the in-memory anonymized output(s) directly into the QA
    # engine without re-uploading. Visible only after an anonymization run
    # produced results in session state. Avoids the download → re-upload
    # round-trip when QA-checking your own anonymized files. The original
    # uploader below still works for files anonymized elsewhere.
    # ------------------------------------------------------------------
    anon_results = st.session_state.get("results") or {}
    if anon_results:
        st.markdown("**🔁 From the current anonymization run**")
        anon_names = list(anon_results.keys())
        if len(anon_names) == 1:
            chosen_name = anon_names[0]
            st.caption(f"Anonymized file in memory: `{chosen_name}`")
        else:
            chosen_name = st.selectbox(
                f"Choose one of the {len(anon_names)} anonymized files",
                options=anon_names,
                key="qa_anon_pick",
            )
        if st.button(
            "🛡️ Check current anonymized output",
            key="qa_use_anon_btn",
            help="Run QA on the in-memory anonymized bytes — no need to download and re-upload.",
        ):
            anon_bytes = anon_results[chosen_name]
            new_sig = (chosen_name, hashlib.sha1(anon_bytes).hexdigest())
            for k in ("qa_results", "qa_target_overrides"):
                st.session_state.pop(k, None)
            for k in [k for k in list(st.session_state.keys()) if k.startswith("qa_edit_")]:
                st.session_state.pop(k, None)
            st.session_state["qa_files"] = [(chosen_name, anon_bytes)]
            st.session_state["qa_filename"] = chosen_name
            st.session_state["qa_original_bytes"] = anon_bytes
            st.session_state["qa_file_signature"] = new_sig
            st.session_state["qa_files_from_uploader"] = False
            st.success(f"Loaded anonymized `{chosen_name}` into QA. Click 🚀 Run QA below.")
            st.rerun()
        st.markdown("<small>Upload a different file</small>", unsafe_allow_html=True)

    qa_uploaded = st.file_uploader(
        "Upload .tmx / .mqxliff files (or one Word .docx)",
        type=["tmx", "mqxliff", "docx"],
        key="qa_file_uploader",
        accept_multiple_files=True,
        help=(
            "QA Check is independent from the Upload tab. Upload several "
            "TMX/MQXLIFF files to check them together — cross-segment checks "
            "(inconsistent translation, date/number formats) then compare "
            "segments ACROSS files too. Word documents are monolingual and "
            "must be checked one at a time."
        ),
    )

    def _qa_clear_loaded_state():
        """Drop every loaded-file artifact so stale results can't be re-run."""
        for k in ("qa_results", "qa_target_overrides", "qa_files",
                  "qa_filename", "qa_original_bytes", "qa_file_signature",
                  "qa_files_from_uploader"):
            st.session_state.pop(k, None)
        for k in [k for k in list(st.session_state.keys()) if k.startswith("qa_edit_")]:
            st.session_state.pop(k, None)

    if qa_uploaded:
        # Multi-file guard: Word documents are monolingual, so they can't be
        # mixed with (or accompany) other files in a combined QA pass.
        docx_names = [f.name for f in qa_uploaded if f.name.lower().endswith(".docx")]
        if docx_names and len(qa_uploaded) > 1:
            st.error(
                "Word documents must be checked one at a time. Remove "
                f"{', '.join(f'`{n}`' for n in docx_names)} or upload it on its own."
            )
            # Invalid batch — clear any previously loaded files/results so
            # Run QA can't silently execute a stale set.
            _qa_clear_loaded_state()
        else:
            # Compare a content hash (sha1 of the bytes) plus the filename for
            # EVERY uploaded file so any change in the payload set — adding,
            # removing or re-uploading a file with different content —
            # refreshes state correctly. Reading bytes once per upload event
            # is cheap relative to the QA pipeline that follows.
            new_files = [(f.name, f.getvalue()) for f in qa_uploaded]
            new_sig = tuple(sorted(
                (name, hashlib.sha1(data).hexdigest()) for name, data in new_files))
            prev_sig = st.session_state.get("qa_file_signature")
            if prev_sig != new_sig:
                for k in ("qa_results", "qa_target_overrides"):
                    st.session_state.pop(k, None)
                # Drop any per-segment edit widget keys from the previous file
                # so the new file starts with empty edit fields.
                for k in [k for k in list(st.session_state.keys()) if k.startswith("qa_edit_")]:
                    st.session_state.pop(k, None)
                st.session_state["qa_files"] = new_files
                st.session_state["qa_filename"] = new_files[0][0]
                st.session_state["qa_original_bytes"] = new_files[0][1]
                st.session_state["qa_file_signature"] = new_sig
                st.session_state["qa_files_from_uploader"] = True
    elif st.session_state.get("qa_files_from_uploader"):
        # The user cleared the uploader — drop the previously uploaded set so
        # Run QA can't re-run stale files. (Files loaded via the in-memory
        # anonymized shortcut are NOT cleared: that path doesn't use the
        # uploader, so an empty uploader is its normal state.)
        _qa_clear_loaded_state()

    # Task #73 — monolingual Word mode. When the loaded file is a Word
    # document there is no source/target pair: the reviewer picks the
    # language (so spell-check uses the right dictionary) and every
    # source↔target comparison check is hidden.
    qa_filename_now = st.session_state.get("qa_filename", "") or ""
    qa_files_now = st.session_state.get("qa_files") or []
    qa_is_word = (len(qa_files_now) <= 1
                  and qa_filename_now.lower().endswith(".docx"))
    if qa_is_word:
        import spellcheck as _sc_langsel
        _lang_opts = sorted(_sc_langsel.DICTIONARY_SOURCES.keys())
        if st.session_state.get("qa_monolingual_lang") not in _lang_opts:
            st.session_state["qa_monolingual_lang"] = (
                "es_ES" if "es_ES" in _lang_opts else _lang_opts[0]
            )
        st.markdown("**📄 Word document (monolingual QA)**")
        st.selectbox(
            "🌐 Document language",
            options=_lang_opts,
            key="qa_monolingual_lang",
            help=(
                "Pick the document's language so spell-check uses the right "
                "dictionary. Only single-text checks run (spelling, "
                "confusables, repeated words, spacing, brackets, forbidden "
                "terms…); source/target comparison checks are skipped."
            ),
        )

    with st.expander("⚙️ Configuration", expanded=False):
        # 35 checks available; 31 active by default. Four opt-IN checks
        # (`qa_checker.OPT_IN_CHECK_IDS`) ship OFF: `polarity_mismatch`
        # (negation flip — too noisy for this user's workflow), `confusable_pairs`
        # (needs user-supplied word groups, high noise) and the two
        # custom-regex checks (Task #68 — the textareas only render once
        # one of those toggles is ticked so the panel stays uncluttered).
        # Task #73 — in monolingual Word mode only the single-text checks
        # (`MONOLINGUAL_CHECK_IDS`) are offered, so the counts shown here
        # must reflect that subset instead of the full bilingual registry.
        if qa_is_word:
            _avail_ids = set(qa_checker.MONOLINGUAL_CHECK_IDS)
        else:
            _avail_ids = set(qa_checker.ALL_CHECK_IDS)
        _n_avail = len(_avail_ids)
        _n_default = len(_avail_ids - qa_checker.OPT_IN_CHECK_IDS)
        cap_col, tick_col, untick_col = st.columns([6, 1, 1])
        with cap_col:
            st.markdown(
                f"**{_n_avail} checks available / {_n_default} active by default.** "
                "Tick or untick as needed."
            )
        # Bulk toggles — mutate every per-check `qa_check_toggle_{cid}`
        # in session_state BEFORE the checkboxes render so the new
        # state is reflected on the same run. Streamlit forbids
        # writing to a widget-managed key after the widget is
        # instantiated, so this has to happen above the per-check
        # loop. `st.rerun()` ensures the user sees the change
        # immediately without needing a second interaction.
        # Word mode offers only the monolingual subset, so the opt-in
        # checks it still exposes are Confusable + Custom forbidden regex
        # (custom_required_regex is bilingual-only and never shown here).
        if qa_is_word:
            _defaults_help = (
                f"Restore the {_n_default} default-ON checks. Leaves the "
                "two opt-in checks (Confusable, Custom forbidden regex) OFF "
                "and resets the glossary sub-options "
                "(case-sensitive, match inflected) to OFF."
            )
        else:
            _defaults_help = (
                f"Restore the {_n_default} default-ON checks. Leaves the "
                "three opt-in checks (Confusable, Custom "
                "forbidden regex, Custom required regex) OFF "
                "and resets the glossary sub-options "
                "(case-sensitive, match inflected) to OFF."
            )
        with tick_col:
            if st.button("✓ Defaults", key="qa_tick_all_btn",
                         help=_defaults_help):
                for _cid in qa_checker.CATEGORY_METADATA:
                    st.session_state[f"qa_check_toggle_{_cid}"] = (
                        _cid not in qa_checker.OPT_IN_CHECK_IDS
                    )
                # Glossary sub-options live outside CATEGORY_METADATA as
                # their own Streamlit checkboxes; reset them too so the
                # "Defaults" button truly restores the day-one state.
                # IMPORTANT: the checkbox widgets own the `_chk` keys —
                # if we only reset the mirror keys, on rerun the widget
                # repopulates the mirrors from its retained value and
                # the box stays ticked.
                st.session_state["qa_glossary_case_chk"] = False
                st.session_state["qa_glossary_case_sensitive"] = False
                st.session_state["qa_glossary_inflected_chk"] = False
                st.session_state["qa_glossary_inflected_forms"] = False
                st.rerun()
        with untick_col:
            if st.button("✗ None", key="qa_untick_all_btn",
                         help="Untick every check at once."):
                for _cid in qa_checker.CATEGORY_METADATA:
                    st.session_state[f"qa_check_toggle_{_cid}"] = False
                st.session_state["qa_glossary_case_chk"] = False
                st.session_state["qa_glossary_case_sensitive"] = False
                st.session_state["qa_glossary_inflected_chk"] = False
                st.session_state["qa_glossary_inflected_forms"] = False
                st.rerun()
        grouped = {}
        for cid, meta in qa_checker.CATEGORY_METADATA.items():
            grouped.setdefault(meta["group"], []).append((cid, meta))
        new_enabled = set()
        # Custom 3-column layout (Task #70). Content has the most
        # checks so it gets the left column alone; Tags has the fewest
        # so it gets the right column alone; the middle column stacks
        # Terminology → Numeric Elements → Spelling so the most
        # commonly tweaked block (Terminology) sits closest to the
        # Inconsistent-translation slider rendered below the columns.
        rendered_cols: Dict[int, List[str]] = {
            0: ["Content"],
            1: ["Terminology", "Numeric Elements", "Spelling"],
            2: ["Tags"],
        }
        # Any future group not listed above falls into the middle column
        # so it never disappears silently.
        _laid_out = {g for col in rendered_cols.values() for g in col}
        for group_name in grouped.keys():
            if group_name not in _laid_out:
                rendered_cols[1].append(group_name)
        cols = st.columns(3)
        for col_idx in (0, 1, 2):
            with cols[col_idx]:
                for group_name in rendered_cols[col_idx]:
                    items = grouped[group_name]
                    # Task #73 — in monolingual Word mode hide every check
                    # that compares source vs target, leaving only the
                    # single-text checks. A group whose checks are all hidden
                    # drops its header too.
                    if qa_is_word:
                        items = [
                            (c, m) for c, m in items
                            if c in qa_checker.MONOLINGUAL_CHECK_IDS
                        ]
                    if not items:
                        continue
                    st.markdown(f"**{items[0][1]['icon']} {group_name}**")
                    for cid, meta in items:
                        key = f"qa_check_toggle_{cid}"
                        if key not in st.session_state:
                            # All checks default ON except the opt-IN trio
                            # (confusable pairs + the two custom-regex checks).
                            st.session_state[key] = (
                                cid not in qa_checker.OPT_IN_CHECK_IDS
                            )
                        checked = st.checkbox(
                            meta['label'],
                            key=key,
                        )
                        if checked:
                            new_enabled.add(cid)
        st.session_state["qa_enabled_checks"] = new_enabled

        # Task #51 — Inconsistent translation threshold slider. Only
        # rendered when the check is enabled, to keep the rest of the
        # config compact. 1.00 = exact source match only (default);
        # below 1.00 enables fuzzy matching via SequenceMatcher, mirroring
        # the Duplicates-tab behavior.
        if "inconsistent_translation" in new_enabled:
            # Constrain the slider to roughly the width of the left
            # column (1/3 of the expander) instead of stretching across
            # the full screen — easier to read and visually anchored to
            # the Content checks above.
            _thr_col, _ = st.columns([1, 2])
            with _thr_col:
                inconsist_thr = st.slider(
                    "Inconsistent translation: source-similarity threshold",
                    min_value=0.80, max_value=1.00,
                    value=float(st.session_state.get("qa_inconsistent_translation_threshold", 1.00)),
                    step=0.01,
                    key="qa_inconsistent_translation_thr",
                    help=(
                        "1.00 = only flag segments whose source is identical. "
                        "Lower values (e.g. 0.90) also flag near-duplicate sources, "
                        "matching the Duplicates tab's fuzzy mode."
                    ),
                )
                st.session_state["qa_inconsistent_translation_threshold"] = inconsist_thr

        # Spell-check sub-config: detected language + ignore-words list. Only
        # rendered when the user has actually enabled the check, to keep the
        # rest of the QA configuration uncluttered.
        if "spellcheck" in new_enabled:
            try:
                import spellcheck as _sc_mod
                supported_count = len(_sc_mod.DICTIONARY_SOURCES)
                if qa_is_word:
                    # Task #73 — Word mode uses the language the reviewer
                    # picked above, not a value auto-detected from the file.
                    _chosen = st.session_state.get("qa_monolingual_lang") or ""
                    st.caption(
                        f"🔤 Spell-check uses the selected document language "
                        f"**{_chosen}**; dictionary downloaded automatically on "
                        f"first use."
                    )
                    raise StopIteration  # skip the bilingual branches below
                detected = (
                    st.session_state.get("qa_results", {}) or {}
                ).get("target_lang") or ""
                norm = _sc_mod.normalize_lang_code(detected) if detected else None
                if detected and norm:
                    st.caption(
                        f"🔤 Detected language: **{norm}** — dictionary "
                        f"downloaded automatically on first use."
                    )
                elif detected and not norm:
                    st.caption(
                        f"⚠️ Detected language `{detected}` not supported "
                        f"({supported_count} languages available). "
                        f"Spell-check skipped."
                    )
                else:
                    st.caption(
                        f"🔤 Language auto-detected from the file; dictionary "
                        f"downloaded automatically on first use "
                        f"({supported_count} languages supported)."
                    )
            except StopIteration:
                pass
            except Exception:
                pass

            ignore_text = st.text_area(
                "Spell-check: ignore words (one per line, proper nouns, brand names, acronyms…)",
                value=st.session_state.get("qa_spellcheck_ignore_text", ""),
                key="qa_spellcheck_ignore_text_area",
                height=90,
            )
            st.session_state["qa_spellcheck_ignore_text"] = ignore_text
            ignore_words = [w.strip() for w in (ignore_text or "").splitlines() if w.strip()]
            st.session_state["qa_spellcheck_ignore"] = ignore_words
            if ignore_words:
                st.caption(f"✅ {len(ignore_words)} word(s) will be ignored by spell-check")

        st.markdown("**📘 Forbidden terms** (one per line, case-insensitive)")
        forbidden_text = st.text_area(
            "Forbidden terms",
            value=st.session_state.get("qa_forbidden_text", ""),
            key="qa_forbidden_text_area",
            height=100,
            label_visibility="collapsed",
        )
        st.session_state["qa_forbidden_text"] = forbidden_text
        forbidden_terms = qa_checker.parse_forbidden_terms(forbidden_text)
        if forbidden_terms:
            st.caption(f"✅ {len(forbidden_terms)} forbidden term(s) loaded")

        # Task #73 — the glossary check (glossary_violation) is bilingual,
        # so its whole sub-panel is hidden in monolingual Word mode.
        glossary_entries = st.session_state.get("qa_glossary", [])
        glossary_case_sensitive = st.session_state.get("qa_glossary_case_sensitive", False)
        glossary_inflected = st.session_state.get("qa_glossary_inflected_forms", False)
        if not qa_is_word:
            st.markdown("**📒 Glossary** (TXT tab-separated, CSV or Excel `.xlsx`. 3 columns: source, target, *optional note*. Header row optional.)")
            glossary_file = st.file_uploader(
                "Glossary file",
                type=["txt", "csv", "xlsx"],
                key="qa_glossary_uploader",
                label_visibility="collapsed",
            )
            # Auto-enable "Match inflected forms" the first time a glossary is
            # uploaded (or when a different glossary file replaces the previous
            # one). Reviewers almost always want inflected matching active when
            # they bother to upload a glossary; the toggle stays user-overridable
            # afterwards because we only nudge the default on a new file event.
            if glossary_file is not None:
                _prev_name = st.session_state.get("qa_glossary_uploaded_name")
                if _prev_name != glossary_file.name:
                    # Streamlit checkbox widgets own their state via their
                    # `key=` once rendered. Setting only the mirror key
                    # (`qa_glossary_inflected_forms`) wouldn't tick the box
                    # on next render, so we seed the widget's own key
                    # (`qa_glossary_inflected_chk`) BEFORE the widget is
                    # instantiated below. Both are kept in sync so the rest
                    # of the QA pipeline (which reads the mirror key) sees
                    # the new value too.
                    st.session_state["qa_glossary_inflected_chk"] = True
                    st.session_state["qa_glossary_inflected_forms"] = True
                    st.session_state["qa_glossary_uploaded_name"] = glossary_file.name
            # Widget keys are the source of truth — passing both `value=` and
            # mutating `st.session_state[key]` elsewhere triggers Streamlit's
            # "default value AND session state" warning. We seed the key once
            # (if missing) and let the widget own it from then on; the auto-tick
            # on a fresh glossary upload writes directly to this same key above.
            st.session_state.setdefault("qa_glossary_case_chk", False)
            st.session_state.setdefault("qa_glossary_inflected_chk", False)
            glossary_case_sensitive = st.checkbox(
                "Glossary case-sensitive matching",
                key="qa_glossary_case_chk",
                help=(
                    "Terms written entirely in UPPERCASE (e.g. 'ON', "
                    "'NOS', 'IL-6') are treated as acronyms and always "
                    "match case-sensitively, even when this is off."
                ),
            )
            st.session_state["qa_glossary_case_sensitive"] = glossary_case_sensitive
            glossary_inflected = st.checkbox(
                "Match inflected forms (Hunspell)",
                key="qa_glossary_inflected_chk",
                help=(
                    "Match inflected forms of single-word entries "
                    "(e.g. 'patient' matches 'patients'). "
                    "Multi-word entries stay literal. Case-insensitive. "
                    "Enter glossary terms in base (dictionary) form — "
                    "e.g. 'patient', not 'patients'."
                ),
            )
            st.session_state["qa_glossary_inflected_forms"] = glossary_inflected
            glossary_entries = []
            if glossary_file is not None:
                glossary_entries = qa_checker.parse_glossary(
                    glossary_file.getvalue(),
                    glossary_file.name,
                    case_sensitive=glossary_case_sensitive,
                )
                st.session_state["qa_glossary"] = glossary_entries
                st.caption(f"✅ {len(glossary_entries)} glossary entries loaded")
            else:
                glossary_entries = st.session_state.get("qa_glossary", [])
                if glossary_entries:
                    st.caption(f"ℹ️ Using previously loaded glossary ({len(glossary_entries)} entries). Re-upload to replace.")

        # ----- Custom regex patterns (Task #61 / #68) -----
        # Both regex checks are opt-IN (default OFF). The intro line, the
        # case-sensitive toggle and the pattern textareas only render when
        # at least one toggle is ON, so the panel stays uncluttered for
        # users who never reach for custom regex. When a toggle is OFF its
        # textarea is skipped and the corresponding pattern list is empty,
        # so previously typed patterns simply don't fire until the user
        # re-enables the check.
        forbidden_on = "custom_forbidden_regex" in new_enabled
        required_on = "custom_required_regex" in new_enabled
        custom_forbidden_patterns: list = []
        custom_required_patterns: list = []
        if forbidden_on or required_on:
            st.markdown(
                "**🧩 Custom regex patterns:** your own Python regular expressions, "
                "evaluated alongside the built-in checks."
            )
            custom_regex_cs = st.checkbox(
                "Custom regex: case-sensitive matching",
                value=st.session_state.get("qa_custom_regex_case_sensitive", False),
                key="qa_custom_regex_cs_chk",
                help="Applies to both textareas below. Default = case-insensitive.",
            )
            st.session_state["qa_custom_regex_case_sensitive"] = custom_regex_cs
        else:
            custom_regex_cs = st.session_state.get(
                "qa_custom_regex_case_sensitive", False
            )

        if forbidden_on:
            custom_forbidden_text = st.text_area(
                "Forbidden patterns (regex, target). Must NOT appear in the target",
                value=st.session_state.get("qa_custom_forbidden_regex", ""),
                key="qa_custom_forbidden_regex_area",
                height=90,
            )
            st.caption(
                "Python regular expressions, one per line. Empty lines and lines "
                "starting with `#` are ignored. Example: `\\bplacebo\\b`."
            )
            st.session_state["qa_custom_forbidden_regex"] = custom_forbidden_text
            custom_forbidden_patterns, custom_forbidden_errors = (
                qa_checker.parse_custom_regex_patterns(
                    custom_forbidden_text, case_sensitive=custom_regex_cs
                )
            )
            for err in custom_forbidden_errors:
                st.error(
                    f"❌ Forbidden pattern line {err['line']} "
                    f"(`{err['raw']}`): {err['error']}"
                )
            if custom_forbidden_patterns:
                st.caption(
                    f"✅ {len(custom_forbidden_patterns)} forbidden regex pattern(s) loaded"
                )

        if required_on:
            custom_required_text = st.text_area(
                "Required patterns (regex, source → target). If matched in source, "
                "must match the same number of times in target",
                value=st.session_state.get("qa_custom_required_regex", ""),
                key="qa_custom_required_regex_area",
                height=90,
            )
            st.caption(
                "Python regular expressions, one per line. Empty lines and lines "
                "starting with `#` are ignored. Example: `\\bNCT\\d{8}\\b` keeps "
                "every clinical-trial ID present in the source."
            )
            st.session_state["qa_custom_required_regex"] = custom_required_text
            custom_required_patterns, custom_required_errors = (
                qa_checker.parse_custom_regex_patterns(
                    custom_required_text, case_sensitive=custom_regex_cs
                )
            )
            for err in custom_required_errors:
                st.error(
                    f"❌ Required pattern line {err['line']} "
                    f"(`{err['raw']}`): {err['error']}"
                )
            if custom_required_patterns:
                st.caption(
                    f"✅ {len(custom_required_patterns)} required regex pattern(s) loaded"
                )

        # ----- Confusable pairs (Task #42) -----
        if "confusable_pairs" in new_enabled:
            try:
                import confusable_pairs as _cp_mod
            except Exception:
                _cp_mod = None
            if _cp_mod is not None:
                st.markdown("**🔤 Confusable pairs** (real-word errors that spell-check can't catch)")
                st.caption(
                    "Add your own confusable word groups via the textarea or "
                    "an .xlsx upload. No built-in language packs are shipped."
                )

                custom_text = st.text_area(
                    "Custom confusable pairs (one group per line, members separated by `|` or `/`)",
                    value=st.session_state.get("qa_confusable_custom_text", ""),
                    key="qa_confusable_custom_area",
                    height=90,
                    help="Examples:\nefectivo|eficaz|eficiente\nprincipal/principle",
                )
                st.session_state["qa_confusable_custom_text"] = custom_text
                custom_pairs = _cp_mod.parse_custom_pairs(custom_text)

                custom_xlsx = st.file_uploader(
                    "…or upload an Excel `.xlsx` with custom pairs (one row per group, one column per member)",
                    type=["xlsx"],
                    key="qa_confusable_custom_uploader",
                )
                if custom_xlsx is not None:
                    xlsx_pairs = _cp_mod.parse_pairs_xlsx(custom_xlsx.getvalue(), custom_xlsx.name)
                    if xlsx_pairs:
                        # Merge avoiding duplicates with textarea entries.
                        seen_keys = {tuple(sorted(m.lower() for m in g)) for g in custom_pairs}
                        for g in xlsx_pairs:
                            k = tuple(sorted(m.lower() for m in g))
                            if k not in seen_keys:
                                seen_keys.add(k)
                                custom_pairs.append(g)
                st.session_state["qa_confusable_custom_pairs"] = custom_pairs
                if custom_pairs:
                    st.caption(f"✅ {len(custom_pairs)} custom confusable group(s) loaded")

    # Show the Run QA button ONLY when a file is actually loaded (either
    # uploaded in this tab or piped in from the Anonymizer). This keeps the
    # initial QA tab clean — no orphan disabled button before there's anything
    # to check.
    can_run = bool(st.session_state.get("qa_files")) or st.session_state.get("qa_original_bytes") is not None
    run_clicked = False
    if can_run:
        _qa_run_l, _qa_run_c, _qa_run_r = st.columns([2, 1, 2])
        with _qa_run_c:
            run_clicked = st.button(
                "🚀 Run QA",
                type="primary",
                key="qa_run_btn",
                use_container_width=True,
            )

    if run_clicked and can_run:
        config = {
            "enabled_checks": st.session_state.get("qa_enabled_checks", qa_checker.DEFAULT_PROFILE),
            "forbidden_terms": forbidden_terms,
            "glossary": st.session_state.get("qa_glossary", []),
            "glossary_case_sensitive": st.session_state.get("qa_glossary_case_sensitive", False),
            "glossary_inflected_forms": st.session_state.get("qa_glossary_inflected_forms", False),
            "inconsistent_translation_threshold": st.session_state.get("qa_inconsistent_translation_threshold", 1.00),
            "spellcheck_ignore": st.session_state.get("qa_spellcheck_ignore", []),
            "confusable_pairs_custom": st.session_state.get("qa_confusable_custom_pairs", []),
            "custom_forbidden_regex_patterns": custom_forbidden_patterns,
            "custom_required_regex_patterns": custom_required_patterns,
            "custom_regex_case_sensitive": custom_regex_cs,
            # Task #73 — monolingual Word mode language (None for bilingual files).
            "monolingual_lang": st.session_state.get("qa_monolingual_lang") if qa_is_word else None,
        }
        try:
            with st.spinner("Running QA checks..."):
                qa_files_run = st.session_state.get("qa_files") or [
                    (st.session_state["qa_filename"],
                     st.session_state["qa_original_bytes"]),
                ]
                results = qa_checker.run_qa_checks_multi(qa_files_run, config)
            st.session_state["qa_results"] = results
            st.session_state["qa_config"] = config
        except ValueError as exc:
            st.error(f"❌ {exc}")
            return
        except Exception as exc:
            st.error(f"❌ Unexpected error during QA: {exc}")
            return

    results = st.session_state.get("qa_results")
    if not results:
        if not can_run:
            st.info("Upload a TMX, MQXLIFF or Word (.docx) file to begin.")
        return

    st.markdown("---")
    summary = results["summary"]
    target_overrides = st.session_state.setdefault("qa_target_overrides", {})

    # ---- Pre-pass: synchronise per-card editor values into the canonical
    # `qa_target_overrides` map BEFORE the summary line is rendered, so the
    # "N segments overridden manually" count reflects the just-typed edit
    # on the very next rerun (rather than lagging by one).
    #
    # Per-card widget keys follow the pattern `qa_edit_{seg_id}__{card_uid}`
    # (see comments around the inner render loop). For each segment that
    # has at least one such key in session_state, we:
    #   1. Look up its canonical value (override entry, or original target).
    #   2. Promote any sibling widget whose value differs from the
    #      canonical (i.e. the user just typed) to the new canonical.
    #   3. Pop ALL sibling widget keys so each per-card editor below
    #      re-initialises this run from `value=canonical_target`.
    seg_originals: dict[str, str] = {}
    for _cat in results["categories"].values():
        for _issue in _cat["issues"]:
            sid = str(_issue["segment_id"])
            if sid not in seg_originals:
                seg_originals[sid] = _issue.get("target") or ""

    sibling_groups: dict[str, list[str]] = {}
    for _k in list(st.session_state.keys()):
        if not _k.startswith("qa_edit_"):
            continue
        rest = _k[len("qa_edit_"):]
        if "__" not in rest:
            continue
        sid = rest.split("__", 1)[0]
        sibling_groups.setdefault(sid, []).append(_k)

    synced_segs_this_run: set = set()
    for sid, keys in sibling_groups.items():
        if sid not in seg_originals:
            # Orphaned widget (e.g. results no longer contain this segment).
            for _k in keys:
                st.session_state.pop(_k, None)
            continue
        original = seg_originals[sid]
        canonical = target_overrides.get(sid, original)
        for _k in keys:
            v = st.session_state.get(_k)
            if v != canonical:
                canonical = v
                break
        if canonical != original:
            target_overrides[sid] = canonical
        else:
            target_overrides.pop(sid, None)
        for _k in keys:
            st.session_state.pop(_k, None)
        synced_segs_this_run.add(sid)

    st.session_state["qa_target_overrides"] = target_overrides
    overrides_pending = len(target_overrides)

    # Task #73 — monolingual Word reports have no source side, so the info
    # line shows a single "Language" field instead of Source · Target.
    if results.get("monolingual"):
        info_line = (
            f"**File:** `{results['filename']}` · **Format:** {results['format'].upper()} · "
            f"**Language:** `{results['target_lang']}` · "
            f"**Segments:** {results['segment_count']}"
        )
    else:
        info_line = (
            f"**File:** `{results['filename']}` · **Format:** {results['format'].upper()} · "
            f"**Source:** `{results['source_lang']}` · **Target:** `{results['target_lang']}` · "
            f"**Segments:** {results['segment_count']}"
        )
    st.markdown(info_line)

    # Top-of-results notices (e.g. "Spell-check skipped: dictionary unavailable…").
    for _notice in results.get("notices", []) or []:
        st.info(_notice)

    if summary["total"] == 0 and overrides_pending == 0:
        st.success("No issues found 🎉. The file passes all enabled checks.")
    else:
        # Hide the manual-override badge when count is 0 — it's noise the
        # rest of the time. Task #68 — single TOTAL pill replaces the
        # previous HIGH / LOW headline badges.
        extras_html = ""
        if overrides_pending > 0:
            extras_html = (
                f" &nbsp; <span style='color:#1a5488;font-weight:600;'>"
                f"{overrides_pending} segments overridden manually</span>"
            )
        st.markdown(
            f"**Summary:** {_qa_total_badge(summary['total'])}{extras_html}",
            unsafe_allow_html=True,
        )

    # Color legend (Task #36): one swatch per QA group, shared between the
    # in-app highlight and the HTML report so users can map a color back to
    # a group at a glance. Hidden when there are no issues to look at.
    if summary["total"] > 0:
        st.markdown(_qa_legend_html(), unsafe_allow_html=True)

    raw_search = st.text_input(
        "🔍 Filter issues by text or segment number",
        value="",
        key="qa_search_input",
        placeholder="Text, or segment numbers: 12 · 12,15,20 · 10-20 · 5, 8-11, 30",
    ).strip()
    # xbench/memoQ-style search options. Segment-number queries bypass these
    # (they're parsed before the text path even runs).
    opt_cols = st.columns(5)
    with opt_cols[0]:
        search_in_source = st.checkbox("Source", value=True, key="qa_search_in_source")
    with opt_cols[1]:
        search_in_target = st.checkbox("Target", value=True, key="qa_search_in_target")
    with opt_cols[2]:
        search_regex = st.checkbox("Regex", value=False, key="qa_search_regex")
    with opt_cols[3]:
        search_whole_words = st.checkbox("Whole words", value=False, key="qa_search_whole_words")
    with opt_cols[4]:
        search_case_sensitive = st.checkbox("Case-sensitive", value=False, key="qa_search_case_sensitive")

    seg_filter = _parse_segment_filter(raw_search)
    search_query = "" if seg_filter is not None else raw_search
    if seg_filter is not None and not seg_filter:
        st.info("No segment numbers parsed from your query.")

    # Compile the search pattern once. None = no active text filter; a
    # compiled pattern = use it; "invalid" sentinel = warn user and skip.
    compiled_pattern = None
    if search_query:
        if not (search_in_source or search_in_target):
            st.info("Tick **Source** and/or **Target** to search inside segments.")
            compiled_pattern = "invalid"
        else:
            flags = 0 if search_case_sensitive else re.IGNORECASE
            pattern_str = search_query if search_regex else re.escape(search_query)
            if search_whole_words:
                pattern_str = r"(?<!\w)" + pattern_str + r"(?!\w)"
            try:
                compiled_pattern = re.compile(pattern_str, flags)
            except re.error as exc:
                st.warning(f"Invalid regex: {exc}")
                compiled_pattern = "invalid"

    # The same segment may appear in several QA categories (e.g. forbidden term +
    # whitespace edges). Every card must show its own editable text_area
    # while sharing the same logical edit value across cards. Streamlit
    # forbids two widgets sharing the same `key` in the same script run, so
    # each card uses a unique per-card widget key
    # (``qa_edit_{seg_id}__{card_uid}``) and we sync them via a single
    # canonical entry in ``qa_target_overrides[seg_id]``.
    #
    # The sync itself was already performed by the pre-pass above (so the
    # summary line's override count is accurate this very rerun). All sibling
    # widget keys for synced segments have been popped, so each per-card
    # editor below initialises from ``value=canonical_target``.
    cat_items = list(results["categories"].items())
    shown_issue_count = 0
    for idx, (cid, cat) in enumerate(cat_items):
        issues = cat["issues"]
        if seg_filter is not None:
            def _seg_id(i):
                v = i.get("segment_id")
                try:
                    return int(v)
                except (TypeError, ValueError):
                    return None
            def _in_filter(i):
                if _seg_id(i) in seg_filter:
                    return True
                # Multi-file runs — also match the per-file segment number
                # the user actually sees on the cards.
                loc = i.get("local_segment_id")
                return loc is not None and loc in seg_filter
            issues = [i for i in issues if _in_filter(i)]
        elif compiled_pattern == "invalid":
            issues = []
        elif compiled_pattern is not None:
            def _matches(i):
                if search_in_source and compiled_pattern.search(i.get("source") or ""):
                    return True
                if search_in_target and compiled_pattern.search(i.get("target") or ""):
                    return True
                return False
            issues = [i for i in issues if _matches(i)]
        if not issues:
            continue
        shown_issue_count += len(issues)
        header = f"{cat['icon']} {cat['label']} ({len(issues)})"
        with st.expander(header, expanded=False):
            # Task #68 — per-category severity badge removed; the group
            # color in the legend / highlighter already conveys the tag.
            visible_issues = issues[:200]
            for issue_idx, issue in enumerate(visible_issues):
                seg_id = str(issue["segment_id"])
                original_target = issue.get("target") or ""
                canonical_target = target_overrides.get(seg_id, original_target)
                is_edited = (canonical_target != original_target)

                edited_badge = (
                    " <span style=\"background:#1a5488;color:#fff;padding:1px 6px;"
                    "border-radius:3px;font-size:0.7rem;font-weight:700;margin-left:6px;\">EDITED</span>"
                    if is_edited else ""
                )
                # Optional reviewer note (glossary column C). Rendered
                # inside the same header card so it visually belongs to
                # the issue, with a softer warm background so it reads
                # as auxiliary context rather than the main message.
                note_text = issue.get("note") or ""
                note_html = (
                    f'<div style="margin-top:6px;padding:5px 10px;background:#fff8e1;'
                    f'border-left:3px solid #f0b429;border-radius:3px;font-size:0.82rem;'
                    f'color:#5e5f6b;line-height:1.35;">'
                    f'<strong>Note:</strong> {_html_escape(note_text)}</div>'
                ) if note_text else ""
                # Multi-file runs — show the real file + per-file segment
                # number; single-file runs keep the plain "Segment N" label.
                if issue.get("file"):
                    _seg_label = (
                        f'📄 {_html_escape(issue["file"])} — Segment '
                        f'{issue.get("local_segment_id") or issue["segment_id"]}'
                    )
                else:
                    _seg_label = f'Segment {issue["segment_id"]}'
                st.markdown(
                    f'<div style="border-left:4px solid #1a5488;padding:8px 14px;margin:14px 0 8px 0;background:#d4dde4;border-radius:4px;">'
                    f'<strong>{_seg_label}:</strong> {_html_escape(issue["message"])}{edited_badge}'
                    f'{note_html}</div>',
                    unsafe_allow_html=True,
                )
                col_s, col_t = st.columns(2)
                # Task #51 — inconsistent_translation paints BOTH cells
                # with the Content group background. Source = plain text
                # (the source is by definition the same / near-same as
                # the reference, so there's nothing meaningful to diff).
                # Target = word-level diff vs the reference target so
                # the reviewer immediately sees what diverged.
                is_inconsist = (
                    cid == "inconsistent_translation"
                    and issue.get("reference_segment_id") is not None
                )
                if is_inconsist:
                    # Task #66 — derive the card chrome from the Content
                    # palette via the single-source-of-truth lookup so the
                    # in-app card matches the exported HTML report (both
                    # render `inconsistent_translation` under Content).
                    _ic_bg, _ic_fg = qa_checker.get_highlight_color(
                        "inconsistent_translation")
                    cell_style = (
                        f"background:{_ic_bg};border:1px solid {_ic_fg};"
                        f"border-left:4px solid {_ic_fg};border-radius:4px;"
                        f"padding:6px 10px;font-size:0.9rem;"
                    )
                else:
                    cell_style = (
                        "background:#fff;border:1px solid #dee2e6;"
                        "border-radius:4px;padding:6px 10px;font-size:0.9rem;"
                    )

                # Task #55 / #66 — stacked "seg X" + "seg Y" block so the
                # reviewer sees both segments together. Top row is the
                # reference (lower seg_id by construction); diff is painted
                # on the cell whose axis diverges. Task #66 dropped the
                # "Ref ·" / "This ·" prefixes — the two-line stack already
                # communicates which row is the reference.
                if is_inconsist:
                    axis = issue.get("mismatch_axis") or "target"
                    ref_src_txt = issue.get("reference_source") or ""
                    ref_tgt_txt = issue.get("reference_target") or ""
                    this_src_txt = issue.get("source") or ""
                    this_tgt_txt = issue.get("target") or ""
                    ref_id = issue.get("reference_segment_id")
                    if axis == "source":
                        ref_src_html_i, this_src_html_i = highlight_diff(ref_src_txt, this_src_txt)
                        ref_tgt_html_i = _html_escape(ref_tgt_txt)
                        this_tgt_html_i = _html_escape(this_tgt_txt)
                    elif axis == "both":
                        ref_src_html_i, this_src_html_i = highlight_diff(ref_src_txt, this_src_txt)
                        ref_tgt_html_i, this_tgt_html_i = highlight_diff(ref_tgt_txt, this_tgt_txt)
                    else:  # "target"
                        ref_src_html_i = _html_escape(ref_src_txt)
                        this_src_html_i = _html_escape(this_src_txt)
                        ref_tgt_html_i, this_tgt_html_i = highlight_diff(ref_tgt_txt, this_tgt_txt)
                    # When the user has edited the segment, the diff vs
                    # the original target no longer maps onto the edited
                    # text — fall back to plain text on the "This" target.
                    if is_edited:
                        this_tgt_html_i = _html_escape(canonical_target)

                    # Multi-file runs — label each stacked row with its real
                    # file + per-file segment number so cross-file
                    # inconsistencies are unambiguous.
                    if issue.get("file"):
                        _ref_lbl = (
                            f'seg {issue.get("reference_local_segment_id") or ref_id} '
                            f'({_html_escape(issue.get("reference_file") or "")})'
                        )
                        _this_lbl = (
                            f'seg {issue.get("local_segment_id") or seg_id} '
                            f'({_html_escape(issue["file"])})'
                        )
                    else:
                        _ref_lbl = f'seg {ref_id}'
                        _this_lbl = f'seg {seg_id}'

                    def _stack_inconsist(ref_html: str, this_html: str) -> str:
                        return (
                            f'<div style="font-size:0.74rem;color:#5e5f6b;'
                            f'font-weight:600;margin-bottom:2px;">{_ref_lbl}</div>'
                            f'<div style="margin-bottom:8px;">{ref_html}</div>'
                            f'<div style="font-size:0.74rem;color:#5e5f6b;'
                            f'font-weight:600;margin-bottom:2px;">{_this_lbl}</div>'
                            f'<div>{this_html}</div>'
                        )

                with col_s:
                    st.markdown("**Source:**")
                    if is_inconsist:
                        source_html = _stack_inconsist(ref_src_html_i, this_src_html_i)
                    else:
                        source_html = _qa_highlight(issue["source"], issue.get("span_source"), cid)
                    # Overlay yellow search highlight on top of the QA
                    # finding highlight, but only on the field(s) the
                    # user is actually searching in.
                    if search_in_source:
                        source_html = _qa_highlight_search(source_html, compiled_pattern)
                    st.markdown(
                        f'<div style="{cell_style}">{source_html}</div>',
                        unsafe_allow_html=True,
                    )
                with col_t:
                    st.markdown("**Target:**")
                    # When the user has edited the segment, the highlight
                    # span computed at QA time no longer maps onto the
                    # edited text — show the canonical edited text
                    # plain-escaped instead.
                    if is_inconsist:
                        target_html = _stack_inconsist(ref_tgt_html_i, this_tgt_html_i)
                    elif is_edited:
                        target_html = _html_escape(canonical_target)
                    else:
                        target_html = _qa_highlight(issue["target"], issue.get("span_target"), cid)
                    if search_in_target:
                        target_html = _qa_highlight_search(target_html, compiled_pattern)
                    st.markdown(
                        f'<div style="{cell_style}">{target_html}</div>',
                        unsafe_allow_html=True,
                    )

                # Visual breathing room between the read-only Source/Target
                # boxes and the editable target widget below.
                st.write("")

                # Per-card editor. Unique widget key per card so multiple
                # cards for the same segment can coexist in one Streamlit
                # script run; their values are kept in lock-step by the
                # sync block above on every rerun.
                card_uid = f"{cid}__{issue_idx}"
                widget_key = f"qa_edit_{seg_id}__{card_uid}"
                edit_cols = st.columns([5, 1])
                with edit_cols[0]:
                    st.text_area(
                        "Edit target (plain text)",
                        value=canonical_target,
                        key=widget_key,
                        height=80,
                        label_visibility="collapsed",
                    )
                with edit_cols[1]:
                    st.button(
                        "↺ Reset",
                        key=f"qa_reset_btn_{seg_id}__{card_uid}",
                        on_click=_qa_reset_segment_edit,
                        args=(seg_id,),
                        disabled=not is_edited,
                        help="Revert this segment to its original target text.",
                    )
                st.caption(
                    "⚠️ Saving an edit replaces the segment as plain text. "
                    "Any inline tags inside this segment will be lost in the downloaded file."
                )

            if len(issues) > 200:
                st.caption(f"… and {len(issues) - 200} more issues in this category (showing first 200)")

    # Empty-state message when an active filter (segment numbers or text
    # search) hides every issue. Without this, the QA results area silently
    # collapses and the user can't tell whether the filter matched nothing
    # or the file truly has no issues.
    if shown_issue_count == 0:
        if seg_filter is not None:
            st.info("🔍 No issues found for the segment number(s) you entered.")
        elif compiled_pattern == "invalid":
            pass  # warning/info already shown above by the search-bar logic
        elif compiled_pattern is not None:
            st.info("🔍 No issues match your search.")

    # `overrides_pending` is already accurate from the pre-pass at the top
    # of this function; per-card editors below cannot mutate it within the
    # same script run (they only enqueue edits for the next rerun's sync).
    if overrides_pending > 0:
        st.button(
            f"↺ Reset all edits ({overrides_pending})",
            key="qa_reset_all_btn",
            on_click=_qa_reset_all_edits,
            help="Revert every manually edited segment to its original target text.",
        )

    st.markdown("---")

    base_name = (st.session_state["qa_filename"] or "file").rsplit(".", 1)
    stem = base_name[0]
    ext = base_name[1] if len(base_name) > 1 else ("tmx" if results["format"] == "tmx" else "mqxliff")
    cleaned_name = f"Cleaned_{stem}.{ext}"
    mime = "application/xml"
    multi_files_meta = results.get("files") or None
    if multi_files_meta:
        stem = "QA_Combined"

    # Task #73 — Word documents are checked monolingually and the cleaned-file
    # download is XML-only (`prepare_qa_download` re-parses the original as
    # MQXLIFF/TMX and would crash on a .docx/.doc). For Word we offer only the
    # HTML report, centered; bilingual files keep the two-button layout.
    is_word_result = bool(results.get("monolingual"))
    if is_word_result:
        _spacer_l, col_rep_html, _spacer_r = st.columns([3, 4, 3])
    else:
        _spacer_l, col_dl, col_rep_html, _spacer_r = st.columns([2, 3, 3, 2])
        with col_dl:
            try:
                if multi_files_meta:
                    # Multi-file run — overrides are keyed by GLOBAL segment
                    # id; split them back per file using each file's offset,
                    # clean every file, and serve one ZIP.
                    import io as _io
                    import zipfile as _zipfile
                    qa_files_dl = dict(st.session_state.get("qa_files") or [])
                    # Fail fast if any analyzed file is no longer loaded —
                    # serving a partial ZIP would silently lose edits.
                    _missing = [_fm["name"] for _fm in multi_files_meta
                                if _fm["name"] not in qa_files_dl]
                    if _missing:
                        raise RuntimeError(
                            "The QA results reference file(s) no longer "
                            f"loaded ({', '.join(_missing)}). Re-upload the "
                            "same set and run QA again.")
                    _zbuf = _io.BytesIO()
                    with _zipfile.ZipFile(_zbuf, "w", _zipfile.ZIP_DEFLATED) as _zf:
                        for _fm in multi_files_meta:
                            _fname = _fm["name"]
                            _off = _fm["offset"]
                            _cnt = _fm["segment_count"]
                            _local_overrides = {}
                            for _gid_str, _txt in target_overrides.items():
                                try:
                                    _gid = int(_gid_str)
                                except (TypeError, ValueError):
                                    continue
                                if _off < _gid <= _off + _cnt:
                                    _local_overrides[str(_gid - _off)] = _txt
                            _src_bytes = qa_files_dl[_fname]
                            _zf.writestr(
                                f"Cleaned_{_fname}",
                                qa_checker.prepare_qa_download(
                                    _src_bytes, _fname,
                                    target_overrides=_local_overrides,
                                ),
                            )
                    cleaned_bytes = _zbuf.getvalue()
                    cleaned_name = "Cleaned_QA_files.zip"
                    mime = "application/zip"
                else:
                    cleaned_bytes = qa_checker.prepare_qa_download(
                        st.session_state["qa_original_bytes"],
                        st.session_state["qa_filename"],
                        target_overrides=target_overrides,
                    )
                if target_overrides:
                    help_text = (
                        f"Applies {len(target_overrides)} inline target edit(s). "
                        "Manual edits replace the target as plain text and discard "
                        "inline tags inside the segment. The original XML structure "
                        "is preserved for every other segment."
                    )
                else:
                    help_text = (
                        "Round trips the original file unchanged. Use the inline "
                        "editor on any issue card to override a target. Those "
                        "edits will be applied here."
                    )
                st.download_button(
                    "📥 Download cleaned files (ZIP)" if multi_files_meta
                    else "📥 Download cleaned file",
                    data=cleaned_bytes,
                    file_name=cleaned_name,
                    mime=mime,
                    key="qa_download_cleaned",
                    disabled=not target_overrides,
                    help=help_text,
                    use_container_width=True,
                )
            except Exception as exc:
                st.error(f"Could not prepare cleaned file: {exc}")
    with col_rep_html:
        try:
            html_bytes = qa_checker.export_qa_report(
                results, "html", target_overrides=target_overrides,
            )
            st.download_button(
                "📰 Export HTML report",
                data=html_bytes,
                file_name=f"QA_Report_{stem}.html",
                mime="text/html",
                key="qa_download_html",
                disabled=summary["total"] == 0,
                use_container_width=True,
            )
        except Exception as exc:
            st.error(f"HTML export failed: {exc}")


# Above this unique-source segment count the variant pass renders a
# real st.progress bar (Task #74) so users can see the analysis
# advancing instead of assuming the app has frozen. Below it the pass
# is near-instant and a flashing bar would just be noise. Kept above
# the sizes used by test_distribution.py so headless test runs never
# touch Streamlit UI elements.
VARIANT_PROGRESS_THRESHOLD = 300


def _collect_dedup_segments(previews: dict, no_anon_segments: dict = None,
                            filter_junk: bool = False, min_words_junk: int = 2,
                            filter_short: bool = False, min_words: int = 5) -> list:
    """Collect filtered, tag-stripped segments shared by all dedup detectors."""
    if not no_anon_segments:
        no_anon_segments = {}
    replacement_token = st.session_state.get('replacement_token', '███')
    token_esc = re.escape(replacement_token)
    consol_pat = rf'{token_esc}(?:\s*[,;.\-–—/|]\s*{token_esc}|\s+{token_esc}|\s*\(\s*{token_esc}\s*\))+'

    all_segments = []
    for filename, file_previews in previews.items():
        for preview in file_previews:
            segment_key = f"{filename}_{preview['segment']}"
            if filter_junk and is_junk_segment(preview, min_words_junk):
                if not st.session_state.get(f"skipjunk_{segment_key}", False):
                    continue
            if no_anon_segments.get(segment_key, False):
                continue
            if filter_short and segment_word_count(preview) < min_words:
                continue

            changed = preview.get('changed', False)
            source_text = preview.get('source_after' if changed else 'source_before', '').strip()
            target_text = preview.get('target_after' if changed else 'target_before', '').strip()
            source_text = strip_inline_tags(source_text)
            target_text = strip_inline_tags(target_text)
            source_text = re.sub(consol_pat, replacement_token, source_text)
            target_text = re.sub(consol_pat, replacement_token, target_text)

            if not source_text and not target_text:
                continue

            all_segments.append({
                "file": filename,
                "segment": preview['segment'],
                "source": source_text,
                "target": target_text,
                "key": segment_key
            })
    return all_segments


def _detect_similar_source_variants(segments: list, source_threshold: int = 85,
                                    target_max: int = 100) -> list:
    """Group segments whose sources are similar but NOT identical (fuzzy).

    ``source_threshold`` (primary) is the minimum source similarity for two
    segments to count as variants. ``target_max`` (secondary, optional) keeps
    a group only when at least one member's target similarity to the reference
    is at most ``target_max`` — lower it to surface variants whose
    translations also differ. ``target_max == 100`` disables the filter.
    """
    groups = []
    if not segments:
        return groups
    thr = source_threshold / 100.0
    tmax = target_max / 100.0
    n = len(segments)

    # ------------------------------------------------------------------
    # Task #74 — exact pruning before the expensive full ratio().
    # difflib guarantees ratio() <= quick_ratio() <= real_quick_ratio(),
    # so any pair whose *upper bound* is already below the threshold can
    # be skipped without changing the result. Two bounds are used:
    #   1. Length bound (== real_quick_ratio): 2*min(la,lb)/(la+lb).
    #      Segments are index-sorted by source length so only the
    #      mathematically reachable length window is visited at all.
    #   2. Character-multiset bound (== quick_ratio): the same integer
    #      `matches` count difflib computes, evaluated for all window
    #      candidates at once on a numpy per-char count matrix. A
    #      pure-Python Counter fallback covers pathological alphabets
    #      (e.g. CJK TMs) where the matrix would be too wide.
    # Survivors pay the same-orientation SequenceMatcher ratio() as
    # before (seq2 prep cached per candidate — set_seq1 does not affect
    # the result), so groups, member order and similarity percentages
    # are EXACTLY the ones the naive O(n²) scan produced — only
    # provably-useless comparisons are skipped.
    # ------------------------------------------------------------------
    sources = [s["source"] for s in segments]
    lens = [len(src) for src in sources]
    counters = [Counter(src) for src in sources]
    alphabet = set()
    for src in sources:
        alphabet.update(src)
    use_np = len(alphabet) <= 4096  # count-matrix width guard
    if use_np:
        col = {ch: k for k, ch in enumerate(sorted(alphabet))}
        mat = np.zeros((n, max(len(alphabet), 1)), dtype=np.int32)
        for idx, cnt in enumerate(counters):
            row = mat[idx]
            for ch, v in cnt.items():
                row[col[ch]] = v
        lens_np = np.array(lens, dtype=np.int64)
        order = np.argsort(lens_np, kind="stable")
        sorted_lens = lens_np[order]
        used_mask = np.zeros(n, dtype=bool)
        used_idx = None
        buckets = bucket_lens = None
    else:
        used_mask = None
        used_idx = set()
        buckets = {}
        for idx, ln in enumerate(lens):
            buckets.setdefault(ln, []).append(idx)
        bucket_lens = sorted(buckets)
    sm_cache = {}  # j -> SequenceMatcher with seq2 = sources[j] prepared

    progress = None
    if n > VARIANT_PROGRESS_THRESHOLD:
        progress = st.progress(
            0.0, text=f"Analyzing similar-source variants… 0 / {n:,} segments"
        )
    report_every = max(1, n // 200)

    for i, seg_a in enumerate(segments):
        if progress is not None and i % report_every == 0:
            progress.progress(
                i / n,
                text=f"Analyzing similar-source variants… {i:,} / {n:,} segments",
            )
        if (used_mask[i] if use_np else i in used_idx):
            continue
        la = lens[i]
        src_a = sources[i]

        # Candidate length window: outside it the length bound alone
        # makes the threshold unreachable. Padded ±1 so borderline
        # float cases stay in; the exact bounds below re-check each one.
        if thr > 0.0:
            lo = int(thr * la / (2.0 - thr)) - 1
            hi = int(la * (2.0 - thr) / thr) + 1
        else:
            lo, hi = 0, max(lens)

        # Group membership does not depend on visit order (`used` only
        # changes AFTER a whole group is accepted), so candidates can be
        # gathered in any order; survivors are re-sorted by original
        # index at the end so the members list keeps the exact same
        # order as the naive j-ascending scan.
        if use_np:
            left = int(np.searchsorted(sorted_lens, lo, side="left"))
            right = int(np.searchsorted(sorted_lens, hi, side="right"))
            j_win = order[left:right]
            j_win = j_win[(j_win > i) & ~used_mask[j_win]]
            survivors = []
            if j_win.size:
                lb = lens_np[j_win]
                total = la + lb
                # total == 0 (two empty sources) bypasses both bounds —
                # difflib returns 1.0 there without dividing. Unreachable
                # with unique sources, kept for faithfulness.
                safe = np.where(total > 0, total, 1)
                # Exact length bound (difflib real_quick_ratio).
                j2 = j_win[(total == 0)
                           | (2.0 * np.minimum(la, lb) / safe >= thr)]
                if j2.size:
                    # Exact character-multiset bound (difflib
                    # quick_ratio): identical integer `matches`.
                    matches = np.minimum(mat[j2], mat[i]).sum(axis=1)
                    total2 = la + lens_np[j2]
                    safe2 = np.where(total2 > 0, total2, 1)
                    survivors = j2[(total2 == 0)
                                   | (2.0 * matches / safe2 >= thr)].tolist()
        else:
            ca = counters[i]
            survivors = []
            start = bisect.bisect_left(bucket_lens, lo)
            for ln in bucket_lens[start:]:
                if ln > hi:
                    break
                total = la + ln
                # Exact length bound (difflib real_quick_ratio) —
                # constant for the whole bucket, checked once.
                if total and 2.0 * (la if la < ln else ln) / total < thr:
                    continue
                lst = buckets[ln]
                for j in lst[bisect.bisect_right(lst, i):]:  # only j > i
                    if j in used_idx:
                        continue
                    # Exact character-multiset bound (difflib
                    # quick_ratio): same integer `matches`; min() is
                    # symmetric so we walk the smaller Counter.
                    if total:
                        cb = counters[j]
                        small, big = (cb, ca) if len(cb) < len(ca) else (ca, cb)
                        bget = big.get
                        matches = 0
                        for ch, va in small.items():
                            vb = bget(ch)
                            if vb:
                                matches += va if va < vb else vb
                        if 2.0 * matches / total < thr:
                            continue
                    survivors.append(j)

        matched = []  # (j, sratio)
        for j in survivors:
            sm = sm_cache.get(j)
            if sm is None:
                sm = difflib.SequenceMatcher()
                sm.set_seq2(sources[j])
                sm_cache[j] = sm
            sm.set_seq1(src_a)
            sratio = sm.ratio()
            if sratio >= thr:
                matched.append((j, sratio))
        if not matched:
            continue
        matched.sort()
        members = [dict(seg_a, _similarity=100.0)]
        members.extend(dict(segments[j], _similarity=r * 100) for j, r in matched)
        if tmax < 1.0:
            differs = any(
                difflib.SequenceMatcher(None, members[0]["target"], m["target"]).ratio() <= tmax
                for m in members[1:]
            )
            if not differs:
                continue
        if use_np:
            used_mask[i] = True
            for j, _ in matched:
                used_mask[j] = True
        else:
            used_idx.add(i)
            used_idx.update(j for j, _ in matched)
        best = max(m["_similarity"] for m in members[1:])
        groups.append({"members": members, "similarity": best})
    if progress is not None:
        progress.empty()
    return groups


def detect_dedup_sections(previews: dict, no_anon_segments: dict = None,
                          filter_junk: bool = False, min_words_junk: int = 2,
                          filter_short: bool = False, min_words: int = 5,
                          variant_source_threshold: int = 85,
                          variant_target_max: int = 100) -> tuple:
    """Partition segments into three disjoint dedup categories.

    Returns ``(identical_groups, conflict_groups, variant_groups)``:
      * identical — same source AND same target (exact redundancy).
      * conflict  — same source, >=2 distinct targets (translation conflicts).
      * variant   — similar (not identical) source, gated by the two variant
        sliders.
    A segment appears in at most one category. Sections 1 and 2 are exact
    (no slider); only section 3 is fuzzy.
    """
    all_segments = _collect_dedup_segments(
        previews, no_anon_segments, filter_junk, min_words_junk,
        filter_short, min_words
    )

    by_source = {}
    for seg in all_segments:
        by_source.setdefault(seg["source"], []).append(seg)

    identical_groups = []
    conflict_groups = []
    singles = []
    for src, segs in by_source.items():
        if len(segs) < 2:
            singles.append(segs[0])
            continue
        distinct_targets = set(s["target"] for s in segs)
        if len(distinct_targets) == 1:
            identical_groups.append({"members": segs, "similarity": 100.0})
        else:
            conflict_groups.append({"members": segs, "similarity": 100.0})

    variant_groups = _detect_similar_source_variants(
        singles, variant_source_threshold, variant_target_max
    )
    return identical_groups, conflict_groups, variant_groups


def _render_dedup_section(groups: list, *, title: str, intro: str,
                          keep_state_key: str, group_keys_state_key: str,
                          radio_prefix: str, expander_label, show_member_sim: bool):
    """Render one Duplicates-tab section (identical / conflict / variant).

    Each section keeps its own keep-choice map in session_state so the clean
    TMX download can honor per-section conserve/discard selections.
    """
    st.markdown(f"### {title}")
    if intro:
        st.markdown(f"<small>{intro}</small>", unsafe_allow_html=True)

    old_keep = st.session_state.get(keep_state_key, {})
    new_keep = {}
    new_group_keys = {}

    for g_idx, group in enumerate(groups):
        members = group["members"]
        similarity = group["similarity"]

        group_keys = sorted([m["key"] for m in members])
        stable_id = "|".join(group_keys)
        new_group_keys[stable_id] = group_keys

        if stable_id in old_keep and (old_keep[stable_id] == "__keep_all__" or old_keep[stable_id] in group_keys):
            new_keep[stable_id] = old_keep[stable_id]
        else:
            new_keep[stable_id] = members[0]["key"]

        with st.expander(expander_label(g_idx, members, similarity), expanded=(g_idx == 0)):
            options = []
            option_keys = []
            for member in members:
                options.append(f"📄 {member['file']} — Segment {member['segment']}")
                option_keys.append(member["key"])
            options.append("Keep all")
            option_keys.append("__keep_all__")

            current_kept = new_keep.get(stable_id, option_keys[0])
            if current_kept not in option_keys:
                current_kept = option_keys[0]
            current_index = option_keys.index(current_kept)

            selected = st.radio(
                "Keep:",
                options=options,
                index=current_index,
                key=f"{radio_prefix}_{stable_id}",
                horizontal=False
            )
            selected_idx = options.index(selected)
            new_keep[stable_id] = option_keys[selected_idx]

            keep_all = new_keep[stable_id] == "__keep_all__"

            ref_source = members[0]["source"][:500]
            ref_target = members[0]["target"][:500]

            for m_idx, member in enumerate(members):
                is_kept = keep_all or (option_keys[m_idx] == new_keep[stable_id])
                if is_kept:
                    badge = '<span style="background:#28a745;color:white;padding:2px 6px;border-radius:4px;font-size:0.75rem;">✓ Kept</span>'
                else:
                    badge = '<span style="background:#dc3545;color:white;padding:2px 6px;border-radius:4px;font-size:0.75rem;">🗑️ Discarded</span>'

                sim_suffix = ""
                if show_member_sim:
                    member_sim = member.get("_similarity", similarity)
                    sim_suffix = f' <small style="color:#666;">({member_sim:.0f}% similar)</small>'

                st.markdown(
                    f'<div style="background:{"#d4edda" if is_kept else "#f8d7da"};border:1px solid {"#28a745" if is_kept else "#dc3545"};border-radius:6px;padding:0.5rem 0.8rem;margin:0.3rem 0;">'
                    f'<strong>📄 {member["file"]} — Segment {member["segment"]}</strong> {badge}{sim_suffix}'
                    f'</div>',
                    unsafe_allow_html=True
                )

                src_text = member["source"][:500]
                tgt_text = member["target"][:500]
                if m_idx == 0:
                    src_html = _html_escape(src_text)
                    tgt_html = _html_escape(tgt_text)
                else:
                    _, src_html = highlight_diff(ref_source, src_text)
                    _, tgt_html = highlight_diff(ref_target, tgt_text)

                col1, col2 = st.columns(2)
                with col1:
                    st.markdown("**Source:**")
                    st.markdown(
                        f'<div style="background:#f8f9fa;border:1px solid #dee2e6;border-radius:4px;padding:0.5rem;font-size:0.9rem;">{src_html}</div>',
                        unsafe_allow_html=True
                    )
                with col2:
                    st.markdown("**Target:**")
                    st.markdown(
                        f'<div style="background:#f8f9fa;border:1px solid #dee2e6;border-radius:4px;padding:0.5rem;font-size:0.9rem;">{tgt_html}</div>',
                        unsafe_allow_html=True
                    )

            st.markdown("---")

    st.session_state[keep_state_key] = new_keep
    st.session_state[group_keys_state_key] = new_group_keys


def extract_existing_canonical_ids(originals: dict) -> dict:
    """Extract existing x-document canonical IDs from original TMX files.
    Returns dict: {(filename, segment_number) -> canonical_id}"""
    from lxml import etree
    canonical_map = {}
    for filename, content in originals.items():
        if not filename.lower().endswith(".tmx"):
            continue
        try:
            if isinstance(content, str):
                content = content.encode('utf-8')
            tree = etree.fromstring(content)
            for i, tu in enumerate(tree.xpath("//tu")):
                doc_prop = tu.find("prop[@type='x-document']")
                if doc_prop is not None and doc_prop.text:
                    canonical_map[(filename, i + 1)] = doc_prop.text.strip()
        except Exception:
            pass
    return canonical_map


def generate_clean_tmx(previews: dict, results: dict, originals: dict,
                       filter_junk: bool, min_words_junk: int,
                       filter_short: bool, min_words: int,
                       exclude_modified: bool, exclusion_threshold: float,
                       excluded_segments: dict, no_anon_segments: dict,
                       dedup_tmx: bool = True,
                       dedup_keep_choices: dict = None,
                       tmx_filename: str = None,
                       canonical_id: str = None,
                       existing_canonical_map: dict = None) -> tuple:
    """Generate a clean TMX containing only valid anonymized segments."""
    from lxml import etree
    
    src_lang = "en"
    tgt_lang = "es"
    
    for filename in originals:
        try:
            tree = etree.fromstring(originals[filename])
            is_tmx = filename.lower().endswith(".tmx")
            if is_tmx:
                header = tree.find(".//header")
                if header is not None:
                    src_lang = header.get("srclang", "en")
                    tus = tree.xpath("//tu/tuv")
                    langs = set()
                    for tuv in tus:
                        lang = tuv.get("{http://www.w3.org/XML/1998/namespace}lang", tuv.get("lang", ""))
                        if lang:
                            langs.add(lang)
                    for lang in langs:
                        if lang.lower() != src_lang.lower():
                            tgt_lang = lang
                            break
            else:
                nsmap = tree.nsmap
                default_ns = nsmap.get(None, '')
                if default_ns:
                    ns = {'x': default_ns}
                    file_els = tree.xpath('//x:file', namespaces=ns)
                else:
                    file_els = tree.xpath('//file')
                if file_els:
                    src_lang = file_els[0].get("source-language", "en")
                    tgt_lang = file_els[0].get("target-language", "es")
            break
        except Exception:
            pass
    
    src_lang = _normalize_lang_code(src_lang)
    tgt_lang = _normalize_lang_code(tgt_lang)
    
    tmx_root = etree.Element("tmx", version="1.4")
    header = etree.SubElement(tmx_root, "header",
                              creationtool="Clean&QA",
                              datatype="PlainText",
                              segtype="sentence",
                              srclang=src_lang)
    has_existing = existing_canonical_map and len(existing_canonical_map) > 0
    if has_existing:
        header_filename = tmx_filename
    else:
        header_filename = canonical_id if canonical_id else tmx_filename
    if header_filename:
        prop = etree.SubElement(header, "prop", type="x-filename")
        prop.text = header_filename
    body = etree.SubElement(tmx_root, "body")
    
    valid_count = 0
    no_anon_skipped = 0
    dedup_count = 0
    dedup_details = []
    excluded_ids = []
    replacement_token = st.session_state.get('replacement_token', '███')

    dedup_discard_keys = set()
    # Task #76 — remember which Duplicates section each discarded key came
    # from so the TM Quality Report can split the dedup count into
    # identical / conflict / variant without a parallel recomputation.
    dedup_discard_sections = {}
    if dedup_keep_choices and dedup_tmx:
        dedup_group_keys = st.session_state.get('dedup_group_keys', {})
        for stable_id, kept_key in dedup_keep_choices.items():
            if kept_key == "__keep_all__":
                continue
            group_keys = dedup_group_keys.get(stable_id, [])
            for k in group_keys:
                if k != kept_key:
                    dedup_discard_keys.add(k)
                    dedup_discard_sections[k] = "identical"

    conflict_keep_choices = st.session_state.get('conflict_keep', {})
    if conflict_keep_choices and dedup_tmx:
        conflict_group_keys = st.session_state.get('conflict_group_keys', {})
        for stable_id, kept_key in conflict_keep_choices.items():
            if kept_key == "__keep_all__":
                continue
            group_keys = conflict_group_keys.get(stable_id, [])
            for k in group_keys:
                if k != kept_key:
                    dedup_discard_keys.add(k)
                    dedup_discard_sections[k] = "conflict"

    variant_keep_choices = st.session_state.get('variant_keep', {})
    if variant_keep_choices and dedup_tmx:
        variant_group_keys = st.session_state.get('variant_group_keys', {})
        for stable_id, kept_key in variant_keep_choices.items():
            if kept_key == "__keep_all__":
                continue
            group_keys = variant_group_keys.get(stable_id, [])
            for k in group_keys:
                if k != kept_key:
                    dedup_discard_keys.add(k)
                    dedup_discard_sections[k] = "variant"
    
    short_excluded = []
    short_anon_excluded = []
    empty_excluded = []
    heavy_excluded = []
    no_anon_excluded = []
    
    for filename, file_previews in previews.items():
        for preview in file_previews:
            segment_key = f"{filename}_{preview['segment']}"
            
            if filter_junk and is_junk_segment(preview, min_words_junk):
                sk = f"skipjunk_{segment_key}"
                if not st.session_state.get(sk, False):
                    excluded_ids.append(preview['segment'])
                    short_excluded.append({
                        "file": filename, "segment": preview['segment'],
                        "source": preview.get('source_before', ''),
                        "target": preview.get('target_before', '')
                    })
                    continue
            
            if no_anon_segments.get(segment_key, False):
                no_anon_skipped += 1
                no_anon_excluded.append({
                    "file": filename, "segment": preview['segment']
                })
                continue
            
            if filter_short and segment_word_count(preview) < min_words:
                sk_short = f"skipshort_{segment_key}"
                if not st.session_state.get(sk_short, False):
                    excluded_ids.append(preview['segment'])
                    short_anon_excluded.append({
                        "file": filename, "segment": preview['segment'],
                        "source_before": preview.get('source_before', ''),
                        "target_before": preview.get('target_before', ''),
                        "source_after": preview.get('source_after', ''),
                        "target_after": preview.get('target_after', '')
                    })
                    continue
            
            changed = preview.get('changed', False)
            source_text = preview.get('source_after' if changed else 'source_before', '').strip()
            target_text = preview.get('target_after' if changed else 'target_before', '').strip()
            
            source_text = strip_inline_tags(source_text)
            target_text = strip_inline_tags(target_text)
            
            token_esc = re.escape(replacement_token)
            consol_pat = rf'{token_esc}(?:\s*[,;.\-–—/|]\s*{token_esc}|\s+{token_esc}|\s*\(\s*{token_esc}\s*\))+'
            source_text = re.sub(consol_pat, replacement_token, source_text)
            target_text = re.sub(consol_pat, replacement_token, target_text)
            
            if not source_text and not target_text:
                excluded_ids.append(preview['segment'])
                empty_excluded.append({
                    "file": filename, "segment": preview['segment']
                })
                continue
            
            if exclude_modified:
                should_exclude = excluded_segments.get(segment_key, None)
                if should_exclude is not False:
                    has_token_src = replacement_token in (preview.get('source_after', ''))
                    has_token_tgt = replacement_token in (preview.get('target_after', ''))
                    if has_token_src or has_token_tgt:
                        src_before = preview.get('source_before', '')
                        tgt_before = preview.get('target_before', '')
                        src_after = preview.get('source_after', '')
                        tgt_after = preview.get('target_after', '')
                        src_pct = (1 - len(src_after.replace(replacement_token, '')) / max(len(src_before), 1)) * 100 if src_before else 0
                        tgt_pct = (1 - len(tgt_after.replace(replacement_token, '')) / max(len(tgt_before), 1)) * 100 if tgt_before else 0
                        max_pct = max(src_pct, tgt_pct)
                        if max_pct >= exclusion_threshold:
                            excluded_ids.append(preview['segment'])
                            heavy_excluded.append({
                                "file": filename, "segment": preview['segment'],
                                "pct": max_pct,
                                "source_after": src_after,
                                "target_after": tgt_after
                            })
                            continue
            
            if dedup_tmx and segment_key in dedup_discard_keys:
                dedup_count += 1
                dedup_details.append({
                    "file": filename, "segment": preview['segment'],
                    "source": source_text, "target": target_text,
                    "similarity": 100.0,
                    "section": dedup_discard_sections.get(segment_key, "identical")
                })
                continue
            
            tu = etree.SubElement(body, "tu")
            
            existing_cid = None
            if existing_canonical_map:
                existing_cid = existing_canonical_map.get((filename, preview['segment']))
            
            if existing_cid:
                doc_prop = etree.SubElement(tu, "prop", type="x-document")
                doc_prop.text = existing_cid
            elif canonical_id:
                doc_prop = etree.SubElement(tu, "prop", type="x-document")
                doc_prop.text = canonical_id
            
            tuv_src = etree.SubElement(tu, "tuv")
            tuv_src.set("{http://www.w3.org/XML/1998/namespace}lang", src_lang)
            seg_src = etree.SubElement(tuv_src, "seg")
            seg_src.text = source_text
            
            tuv_tgt = etree.SubElement(tu, "tuv")
            tuv_tgt.set("{http://www.w3.org/XML/1998/namespace}lang", tgt_lang)
            seg_tgt = etree.SubElement(tuv_tgt, "seg")
            seg_tgt.text = target_text
            
            valid_count += 1
    
    exclusion_breakdown = {
        "short": short_excluded,
        "short_anon": short_anon_excluded,
        "empty": empty_excluded,
        "heavy": heavy_excluded,
        "dedup": dedup_details,
        "no_anon": no_anon_excluded
    }
    
    result = b'<?xml version="1.0" encoding="UTF-8"?>\n'
    result += etree.tostring(tmx_root, encoding="unicode", pretty_print=True).encode("utf-8")
    
    return result, valid_count, no_anon_skipped, excluded_ids, dedup_count, dedup_details, exclusion_breakdown


ANON_CATEGORIES = [
    ("safe_regex", "Safe Regex"),
    ("regex_ct", "Clinical ID Regex"),
    ("proper_names", "Proper Names"),
    ("dictionary", "Custom dictionary"),
]

QA_GROUP_ORDER = ["Content", "Numeric Elements", "Tags", "Terminology", "Spelling"]


def collect_quality_metrics(previews: dict, all_stats: dict,
                            exclusion_breakdown: dict,
                            qa_results: dict = None) -> dict:
    """Task #76 — consolidate session quality metrics per file + batch.

    All counts come straight from the SAME structures the clean-TMX
    generation produced (`exclusion_breakdown` from generate_clean_tmx),
    so the funnel always matches the downloaded TMX exactly. Session-only:
    nothing is persisted.
    """
    reasons = ["short", "short_anon", "empty", "heavy", "no_anon"]
    files = {}
    for fname, file_previews in previews.items():
        stats = all_stats.get(fname, {})
        files[fname] = {
            "raw": len(file_previews),
            "short": 0, "short_anon": 0, "empty": 0, "heavy": 0,
            "no_anon": 0,
            "dedup_identical": 0, "dedup_conflict": 0, "dedup_variant": 0,
            "anon": {key: stats.get(key, 0) for key, _ in ANON_CATEGORIES},
            "unique_terms": stats.get("unique_terms"),
            "wall_ms": stats.get("wall_ms"),
        }
        files[fname]["anon_total"] = sum(files[fname]["anon"].values())

    eb = exclusion_breakdown or {}
    for reason in reasons:
        for item in eb.get(reason, []):
            f = files.get(item.get("file"))
            if f is not None:
                f[reason] += 1
    for item in eb.get("dedup", []):
        f = files.get(item.get("file"))
        if f is not None:
            section = item.get("section", "identical")
            if section not in ("identical", "conflict", "variant"):
                section = "identical"
            f["dedup_" + section] += 1

    for f in files.values():
        f["dedup_total"] = (f["dedup_identical"] + f["dedup_conflict"]
                            + f["dedup_variant"])
        f["excluded_total"] = (f["short"] + f["short_anon"] + f["empty"]
                               + f["heavy"] + f["no_anon"] + f["dedup_total"])
        f["final"] = f["raw"] - f["excluded_total"]
        f["retention_pct"] = (f["final"] / f["raw"] * 100.0) if f["raw"] else 0.0
        if f["wall_ms"] and f["raw"]:
            f["segs_per_s"] = f["raw"] / (f["wall_ms"] / 1000.0)
        else:
            f["segs_per_s"] = None

    batch = {
        "raw": sum(f["raw"] for f in files.values()),
        "final": sum(f["final"] for f in files.values()),
        "anon_total": sum(f["anon_total"] for f in files.values()),
        "unique_terms": sum(f["unique_terms"] or 0 for f in files.values()),
        "anon": {key: sum(f["anon"][key] for f in files.values())
                 for key, _ in ANON_CATEGORIES},
    }
    for reason in reasons + ["dedup_identical", "dedup_conflict",
                             "dedup_variant", "dedup_total", "excluded_total"]:
        batch[reason] = sum(f[reason] for f in files.values())
    batch["retention_pct"] = (batch["final"] / batch["raw"] * 100.0) if batch["raw"] else 0.0
    wall_total = sum(f["wall_ms"] or 0 for f in files.values())
    batch["wall_ms"] = wall_total or None
    batch["segs_per_s"] = (batch["raw"] / (wall_total / 1000.0)) if wall_total else None

    qa = {"run": False, "filename": None, "total": 0, "by_group": {}}
    if qa_results:
        qa["run"] = True
        qa["filename"] = qa_results.get("filename")
        qa["total"] = qa_results.get("summary", {}).get("total", 0)
        by_group = {g: 0 for g in QA_GROUP_ORDER}
        for cat in qa_results.get("categories", {}).values():
            by_group[cat.get("group", "Content")] = (
                by_group.get(cat.get("group", "Content"), 0) + cat.get("count", 0))
        qa["by_group"] = by_group

    config = {
        "Replacement token": st.session_state.get('replacement_token', '███'),
        "Exclude short segments": st.session_state.get('filter_junk', False),
        "Minimum words (short segments)": st.session_state.get('min_words_junk', 2),
        "Exclude short anonymized segments": st.session_state.get('filter_short_segments', False),
        "Minimum words (anonymized segments)": st.session_state.get('min_words', 5),
        "Exclude heavily anonymized": st.session_state.get('exclude_modified_targets', False),
        "Redaction threshold (>=%)": st.session_state.get('exclusion_threshold', 20),
        "Deduplicate TMX segments": st.session_state.get('dedup_tmx', True),
        "Variant source similarity (>=%)": st.session_state.get('variant_source_threshold', 85),
        "Variant max. target similarity (<=%)": st.session_state.get('variant_target_max', 100),
    }
    layers = st.session_state.get('session_layers') or {}

    return {"files": files, "batch": batch, "qa": qa,
            "config": config, "layers": layers}


def generate_changes_excel(dedup_details: list = None, exclusion_breakdown: dict = None,
                           file_canonical_map: dict = None,
                           quality_metrics: dict = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Anonymization Report"
    
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1a5488", end_color="1a5488", fill_type="solid")
    title_font = Font(bold=True, size=14, color="1a5488")
    border = Border(
        left=Side(style='thin', color='bcbdbe'),
        right=Side(style='thin', color='bcbdbe'),
        top=Side(style='thin', color='bcbdbe'),
        bottom=Side(style='thin', color='bcbdbe')
    )
    alt_fill = PatternFill(start_color="e8f4fc", end_color="e8f4fc", fill_type="solid")
    total_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
    total_font = Font(bold=True, size=11)
    section_fill = PatternFill(start_color="d6e4f0", end_color="d6e4f0", fill_type="solid")
    section_font = Font(bold=True, size=12, color="1a5488")
    
    ws.merge_cells('A1:H1')
    ws['A1'] = "ANONYMIZATION REPORT - MQXLIFF/TMX Clean&QA v6.3"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    row = 3
    if file_canonical_map:
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = "SOURCE FILES & CANONICAL IDs"
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        ws[f'B{row}'].fill = section_fill
        ws[f'C{row}'].fill = section_fill
        row += 1
        for fname, cids in file_canonical_map.items():
            ws.cell(row=row, column=1, value=fname).border = border
            if cids == "None":
                ws.merge_cells(f'B{row}:C{row}')
            ws.cell(row=row, column=2, value=cids).border = border
            ws.cell(row=row, column=3).border = border
            row += 1
        row += 1
    
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "STATISTICS SUMMARY"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    
    row += 1
    stats_headers = ["File", "Safe Regex", "Regex CT IDs", "Proper Names", "Dictionary", "Total"]
    for col, header in enumerate(stats_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    row += 1
    grand_total = 0
    for filename, stats in st.session_state.all_stats.items():
        file_total = stats.get("safe_regex", 0) + stats.get("regex_ct", 0) + stats.get("proper_names", 0) + stats.get("dictionary", 0)
        grand_total += file_total
        data = [filename, stats.get("safe_regex", 0), stats["regex_ct"], stats.get("proper_names", 0), stats["dictionary"], file_total]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col > 1:
                cell.alignment = Alignment(horizontal='center')
        row += 1
    
    total_data = ["GRAND TOTAL", "", "", "", "", "", "", grand_total]
    for col, value in enumerate(total_data, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = border
        if col == 1:
            cell.alignment = Alignment(horizontal='left')
        else:
            cell.alignment = Alignment(horizontal='center')
    
    row += 3
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "CHANGES DETAIL"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    
    row += 1
    detail_headers = ["File", "Segment", "Type", "Original Text", "Anonymized Text"]
    for col, header in enumerate(detail_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    row += 1
    alt_row = False
    filter_junk = st.session_state.get('filter_junk', False)
    min_words_junk = st.session_state.get('min_words_junk', 2)
    filter_short = st.session_state.get('filter_short_segments', False)
    min_words = st.session_state.get('min_words', 5)
    
    for filename, file_previews in st.session_state.previews.items():
        for preview in file_previews:
            if filter_junk and is_junk_segment(preview, min_words_junk):
                continue
            if not preview.get('changed', True):
                continue
            if filter_short and segment_word_count(preview) < min_words:
                continue
            
            if preview['source_before'] != preview['source_after']:
                data = [filename, preview['segment'], "Source", preview['source_before'], preview['source_after']]
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    if alt_row:
                        cell.fill = alt_fill
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                row += 1
                alt_row = not alt_row
            if preview['target_before'] != preview['target_after']:
                data = [filename, preview['segment'], "Target", preview['target_before'], preview['target_after']]
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    if alt_row:
                        cell.fill = alt_fill
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                row += 1
                alt_row = not alt_row
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 55
    ws.column_dimensions['E'].width = 55
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 10
    
    ws2 = wb.create_sheet(title="Filtered Segments")
    
    ws2.merge_cells('A1:F1')
    ws2['A1'] = "FILTERED SEGMENTS - Clean&QA v6.3"
    ws2['A1'].font = title_font
    ws2['A1'].alignment = Alignment(horizontal='center')
    
    ex_row = 3
    
    if exclusion_breakdown is None:
        exclusion_breakdown = {"short": [], "short_anon": [], "empty": [], "heavy": [], "dedup": []}
    
    bd_short = exclusion_breakdown.get("short", [])
    bd_short_anon = exclusion_breakdown.get("short_anon", [])
    bd_empty = exclusion_breakdown.get("empty", [])
    bd_heavy = exclusion_breakdown.get("heavy", [])
    bd_dedup = exclusion_breakdown.get("dedup", [])
    
    junk_fill = PatternFill(start_color="e2e3e5", end_color="e2e3e5", fill_type="solid")
    short_fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
    empty_fill = PatternFill(start_color="d6d8db", end_color="d6d8db", fill_type="solid")
    tm_fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
    dedup_fill = PatternFill(start_color="d1ecf1", end_color="d1ecf1", fill_type="solid")
    
    threshold = st.session_state.get('exclusion_threshold', 20)
    has_any = len(bd_short) + len(bd_short_anon) + len(bd_empty) + len(bd_heavy) + len(bd_dedup) > 0
    
    if bd_short:
        ws2.merge_cells(f'A{ex_row}:D{ex_row}')
        ws2[f'A{ex_row}'] = f"SHORT SEGMENTS - Excluded from TM (<{min_words_junk} words or only numbers/symbols)"
        ws2[f'A{ex_row}'].font = Font(bold=True, size=12, color="495057")
        ws2[f'A{ex_row}'].fill = junk_fill
        ex_row += 1
        for col, header in enumerate(["File", "Segment", "Source (original)", "Target (original)"], 1):
            cell = ws2.cell(row=ex_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        ex_row += 1
        for item in bd_short:
            for col, value in enumerate([item["file"], item["segment"], item["source"], item["target"]], 1):
                cell = ws2.cell(row=ex_row, column=col, value=value)
                cell.border = border
                cell.fill = PatternFill(start_color="f0f0f0", end_color="f0f0f0", fill_type="solid")
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            ex_row += 1
        ws2.merge_cells(f'A{ex_row}:D{ex_row}')
        ws2[f'A{ex_row}'] = f"Total: {len(bd_short)} short segments excluded from TM"
        ws2[f'A{ex_row}'].font = total_font
        ws2[f'A{ex_row}'].fill = junk_fill
        ex_row += 1
        ex_row += 2
    
    if bd_short_anon:
        ws2.merge_cells(f'A{ex_row}:F{ex_row}')
        ws2[f'A{ex_row}'] = f"SHORT ANONYMIZED SEGMENTS - Excluded from TM (less than {min_words} words)"
        ws2[f'A{ex_row}'].font = Font(bold=True, size=12, color="856404")
        ws2[f'A{ex_row}'].fill = short_fill
        ex_row += 1
        for col, header in enumerate(["File", "Segment", "Source (original)", "Target (original)", "Source (anonymized)", "Target (anonymized)"], 1):
            cell = ws2.cell(row=ex_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        ex_row += 1
        for item in bd_short_anon:
            for col, value in enumerate([item["file"], item["segment"], item["source_before"], item["target_before"], item["source_after"], item["target_after"]], 1):
                cell = ws2.cell(row=ex_row, column=col, value=value)
                cell.border = border
                cell.fill = PatternFill(start_color="fffbe6", end_color="fffbe6", fill_type="solid")
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            ex_row += 1
        ws2.merge_cells(f'A{ex_row}:F{ex_row}')
        ws2[f'A{ex_row}'] = f"Total: {len(bd_short_anon)} short anonymized segments excluded from TM"
        ws2[f'A{ex_row}'].font = total_font
        ws2[f'A{ex_row}'].fill = short_fill
        ex_row += 1
        ex_row += 2
    
    if bd_empty:
        ws2.merge_cells(f'A{ex_row}:B{ex_row}')
        ws2[f'A{ex_row}'] = "EMPTY SEGMENTS - Excluded from TM (no text content)"
        ws2[f'A{ex_row}'].font = Font(bold=True, size=12, color="495057")
        ws2[f'A{ex_row}'].fill = empty_fill
        ex_row += 1
        for col, header in enumerate(["File", "Segment"], 1):
            cell = ws2.cell(row=ex_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        ex_row += 1
        for item in bd_empty:
            ws2.cell(row=ex_row, column=1, value=item["file"]).border = border
            ws2.cell(row=ex_row, column=2, value=item["segment"]).border = border
            ws2.cell(row=ex_row, column=2).alignment = Alignment(horizontal='center')
            ex_row += 1
        ws2.merge_cells(f'A{ex_row}:B{ex_row}')
        ws2[f'A{ex_row}'] = f"Total: {len(bd_empty)} empty segments excluded from TM"
        ws2[f'A{ex_row}'].font = total_font
        ws2[f'A{ex_row}'].fill = empty_fill
        ex_row += 1
        ex_row += 2
    
    if bd_heavy:
        ws2.merge_cells(f'A{ex_row}:E{ex_row}')
        ws2[f'A{ex_row}'] = f"HEAVILY ANONYMIZED SEGMENTS (redaction >= {threshold}%)"
        ws2[f'A{ex_row}'].font = Font(bold=True, size=12, color="721c24")
        ws2[f'A{ex_row}'].fill = tm_fill
        ex_row += 1
        for col, header in enumerate(["File", "Segment", "Redaction %", "Source (anonymized)", "Target (anonymized)"], 1):
            cell = ws2.cell(row=ex_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        ex_row += 1
        for item in bd_heavy:
            for col, value in enumerate([item["file"], item["segment"], f"{item['pct']:.1f}%", item["source_after"], item["target_after"]], 1):
                cell = ws2.cell(row=ex_row, column=col, value=value)
                cell.border = border
                cell.fill = PatternFill(start_color="fce4ec", end_color="fce4ec", fill_type="solid")
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            ex_row += 1
        ws2.merge_cells(f'A{ex_row}:E{ex_row}')
        ws2[f'A{ex_row}'] = f"Total: {len(bd_heavy)} heavily anonymized segments excluded from TM"
        ws2[f'A{ex_row}'].font = total_font
        ws2[f'A{ex_row}'].fill = tm_fill
        ex_row += 1
        ex_row += 2
    
    if bd_dedup:
        ws2.merge_cells(f'A{ex_row}:D{ex_row}')
        ws2[f'A{ex_row}'] = "DUPLICATE / CONFLICT / VARIANT SEGMENTS - Excluded from TMX (per keep choices)"
        ws2[f'A{ex_row}'].font = Font(bold=True, size=12, color="0c5460")
        ws2[f'A{ex_row}'].fill = dedup_fill
        ex_row += 1
        for col, header in enumerate(["File", "Segment", "Source", "Target"], 1):
            cell = ws2.cell(row=ex_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        ex_row += 1
        for dd in bd_dedup:
            for col, value in enumerate([dd['file'], dd['segment'], dd['source'], dd['target']], 1):
                cell = ws2.cell(row=ex_row, column=col, value=value)
                cell.border = border
                cell.fill = PatternFill(start_color="e8f6f8", end_color="e8f6f8", fill_type="solid")
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            ex_row += 1
        ws2.merge_cells(f'A{ex_row}:D{ex_row}')
        ws2[f'A{ex_row}'] = f"Total: {len(bd_dedup)} segments excluded from TMX"
        ws2[f'A{ex_row}'].font = total_font
        ws2[f'A{ex_row}'].fill = dedup_fill
        ex_row += 1
    
    if not has_any:
        ws2.cell(row=ex_row, column=1, value="No segments were excluded.").font = Font(italic=True, color="666666")
    
    ws2.column_dimensions['A'].width = 40
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 20
    ws2.column_dimensions['D'].width = 45
    ws2.column_dimensions['E'].width = 45
    ws2.column_dimensions['F'].width = 20
    
    # ── Task #76 — "TM Quality Report" summary sheet, inserted FIRST ──
    if quality_metrics:
        qm = quality_metrics
        ws0 = wb.create_sheet(title="TM Quality Report", index=0)
        wb.active = 0

        ws0.merge_cells('A1:H1')
        ws0['A1'] = "TM QUALITY REPORT - MQXLIFF/TMX Clean&QA v6.3 (current session)"
        ws0['A1'].font = title_font
        ws0['A1'].alignment = Alignment(horizontal='center')

        b = qm["batch"]
        r = 3
        ws0.merge_cells(f'A{r}:H{r}')
        ws0[f'A{r}'] = (f"RETENTION RATE: {b['retention_pct']:.1f}% — "
                        f"{b['final']} of {b['raw']} raw segments entered the final clean TM")
        ws0[f'A{r}'].font = Font(bold=True, size=13, color="1a5488")
        ws0[f'A{r}'].fill = total_fill
        r += 2

        # Funnel (batch)
        ws0.merge_cells(f'A{r}:C{r}')
        ws0[f'A{r}'] = "CLEANING FUNNEL (batch)"
        ws0[f'A{r}'].font = section_font
        ws0[f'A{r}'].fill = section_fill
        r += 1
        for col, header in enumerate(["Stage", "Segments", "% of raw"], 1):
            cell = ws0.cell(row=r, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        r += 1
        raw = b["raw"] or 1
        funnel_rows = [
            ("Raw segments", b["raw"], None),
            ("Discarded: junk / short segments", b["short"], b["short"] / raw),
            ("Discarded: short after anonymization", b["short_anon"], b["short_anon"] / raw),
            ("Discarded: empty after cleaning", b["empty"], b["empty"] / raw),
            ("Discarded: heavily anonymized", b["heavy"], b["heavy"] / raw),
            ("Discarded: kept without anonymization", b["no_anon"], b["no_anon"] / raw),
            ("Discarded: duplicates (all sections)", b["dedup_total"], b["dedup_total"] / raw),
            ("Final segments in clean TM", b["final"], b["final"] / raw),
        ]
        alt = False
        for label, count, pct in funnel_rows:
            is_final = label.startswith("Final")
            for col, value in enumerate(
                    [label, count, f"{pct * 100:.1f}%" if pct is not None else "100.0%"], 1):
                cell = ws0.cell(row=r, column=col, value=value)
                cell.border = border
                if is_final:
                    cell.fill = total_fill
                    cell.font = total_font
                elif alt:
                    cell.fill = alt_fill
                if col > 1:
                    cell.alignment = Alignment(horizontal='center')
            alt = not alt
            r += 1
        r += 1

        # Per-file table
        ws0.merge_cells(f'A{r}:J{r}')
        ws0[f'A{r}'] = "PER-FILE BREAKDOWN"
        ws0[f'A{r}'].font = section_font
        ws0[f'A{r}'].fill = section_fill
        r += 1
        pf_headers = ["File", "Raw", "Junk/short", "Short anon.", "Empty",
                      "Heavy anon.", "No-anon kept", "Duplicates", "Final", "Retention %"]
        for col, header in enumerate(pf_headers, 1):
            cell = ws0.cell(row=r, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        r += 1
        alt = False
        for fname, f in qm["files"].items():
            data = [fname, f["raw"], f["short"], f["short_anon"], f["empty"],
                    f["heavy"], f["no_anon"], f["dedup_total"], f["final"],
                    f"{f['retention_pct']:.1f}%"]
            for col, value in enumerate(data, 1):
                cell = ws0.cell(row=r, column=col, value=value)
                cell.border = border
                if alt:
                    cell.fill = alt_fill
                if col > 1:
                    cell.alignment = Alignment(horizontal='center')
            alt = not alt
            r += 1
        r += 1

        # Duplicates breakdown
        ws0.merge_cells(f'A{r}:B{r}')
        ws0[f'A{r}'] = "DUPLICATES BREAKDOWN (discarded from clean TM)"
        ws0[f'A{r}'].font = section_font
        ws0[f'A{r}'].fill = section_fill
        r += 1
        for col, header in enumerate(["Section", "Seg. discarded"], 1):
            cell = ws0.cell(row=r, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        r += 1
        for label, count in [("Identical (same source & target)", b["dedup_identical"]),
                             ("Same-source conflicts resolved", b["dedup_conflict"]),
                             ("Similar-source variants discarded", b["dedup_variant"]),
                             ("Total", b["dedup_total"])]:
            cell = ws0.cell(row=r, column=1, value=label)
            cell.border = border
            cell2 = ws0.cell(row=r, column=2, value=count)
            cell2.border = border
            cell2.alignment = Alignment(horizontal='center')
            if label == "Total":
                cell.font = total_font
                cell.fill = total_fill
                cell2.font = total_font
                cell2.fill = total_fill
            r += 1
        r += 1

        # Anonymization
        ws0.merge_cells(f'A{r}:B{r}')
        ws0[f'A{r}'] = "SENSITIVE DATA REMOVED (anonymization by layer)"
        ws0[f'A{r}'].font = section_font
        ws0[f'A{r}'].fill = section_fill
        r += 1
        for col, header in enumerate(["Layer", "Replacements"], 1):
            cell = ws0.cell(row=r, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        r += 1
        for key, label in ANON_CATEGORIES:
            cell = ws0.cell(row=r, column=1, value=label)
            cell.border = border
            cell2 = ws0.cell(row=r, column=2, value=b["anon"].get(key, 0))
            cell2.border = border
            cell2.alignment = Alignment(horizontal='center')
            r += 1
        for label, value in [("Total replacements", b["anon_total"]),
                             ("Unique terms anonymized (sum per file)", b["unique_terms"])]:
            cell = ws0.cell(row=r, column=1, value=label)
            cell.font = total_font
            cell.fill = total_fill
            cell.border = border
            cell2 = ws0.cell(row=r, column=2, value=value)
            cell2.font = total_font
            cell2.fill = total_fill
            cell2.border = border
            cell2.alignment = Alignment(horizontal='center')
            r += 1
        r += 1

        # QA summary
        ws0.merge_cells(f'A{r}:B{r}')
        ws0[f'A{r}'] = "QA CHECK SUMMARY"
        ws0[f'A{r}'].font = section_font
        ws0[f'A{r}'].fill = section_fill
        r += 1
        qa = qm["qa"]
        if qa["run"]:
            for col, header in enumerate(["Group", "Issues found"], 1):
                cell = ws0.cell(row=r, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            r += 1
            for group in QA_GROUP_ORDER:
                cell = ws0.cell(row=r, column=1, value=group)
                cell.border = border
                cell2 = ws0.cell(row=r, column=2, value=qa["by_group"].get(group, 0))
                cell2.border = border
                cell2.alignment = Alignment(horizontal='center')
                r += 1
            cell = ws0.cell(row=r, column=1, value=f"Total ({qa['filename']})")
            cell.font = total_font
            cell.fill = total_fill
            cell.border = border
            cell2 = ws0.cell(row=r, column=2, value=qa["total"])
            cell2.font = total_font
            cell2.fill = total_fill
            cell2.border = border
            cell2.alignment = Alignment(horizontal='center')
            r += 1
        else:
            ws0.cell(row=r, column=1, value="QA not run in this session.").font = Font(italic=True, color="666666")
            r += 1
        r += 1

        # Processing speed
        if b.get("wall_ms"):
            ws0.merge_cells(f'A{r}:C{r}')
            ws0[f'A{r}'] = "PROCESSING SPEED"
            ws0[f'A{r}'].font = section_font
            ws0[f'A{r}'].fill = section_fill
            r += 1
            for col, header in enumerate(["File", "Time (s)", "Segments/s"], 1):
                cell = ws0.cell(row=r, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            r += 1
            for fname, f in qm["files"].items():
                if not f.get("wall_ms"):
                    continue
                data = [fname, f"{f['wall_ms'] / 1000.0:.1f}",
                        f"{f['segs_per_s']:.1f}" if f.get("segs_per_s") else "—"]
                for col, value in enumerate(data, 1):
                    cell = ws0.cell(row=r, column=col, value=value)
                    cell.border = border
                    if col > 1:
                        cell.alignment = Alignment(horizontal='center')
                r += 1
            data = ["Batch total", f"{b['wall_ms'] / 1000.0:.1f}",
                    f"{b['segs_per_s']:.1f}" if b.get("segs_per_s") else "—"]
            for col, value in enumerate(data, 1):
                cell = ws0.cell(row=r, column=col, value=value)
                cell.font = total_font
                cell.fill = total_fill
                cell.border = border
                if col > 1:
                    cell.alignment = Alignment(horizontal='center')
            r += 2

        # Session configuration snapshot
        ws0.merge_cells(f'A{r}:B{r}')
        ws0[f'A{r}'] = "SESSION CONFIGURATION"
        ws0[f'A{r}'].font = section_font
        ws0[f'A{r}'].fill = section_fill
        r += 1
        for col, header in enumerate(["Setting", "Value"], 1):
            cell = ws0.cell(row=r, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        r += 1
        for label, on in (qm.get("layers") or {}).items():
            cell = ws0.cell(row=r, column=1, value=f"Layer: {label}")
            cell.border = border
            cell2 = ws0.cell(row=r, column=2, value="ON" if on else "OFF")
            cell2.border = border
            cell2.alignment = Alignment(horizontal='center')
            r += 1
        for label, value in qm["config"].items():
            cell = ws0.cell(row=r, column=1, value=label)
            cell.border = border
            if isinstance(value, bool):
                value = "ON" if value else "OFF"
            cell2 = ws0.cell(row=r, column=2, value=value)
            cell2.border = border
            cell2.alignment = Alignment(horizontal='center')
            r += 1

        ws0.column_dimensions['A'].width = 42
        ws0.column_dimensions['B'].width = 14
        for c in "CDEFGH":
            ws0.column_dimensions[c].width = 13
        ws0.column_dimensions['I'].width = 10
        ws0.column_dimensions['J'].width = 13

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@st.dialog("Help & Quick Guide", width="large")
def _render_help_dialog():
    """Visual, plain-English summary of README / USER_GUIDE / VERSION_HISTORY."""
    st.markdown(
        """
<div style="color:#130e45; font-size:0.95rem;">

<p style="color:#5e5f6b; line-height:1.6; margin:0 0 1.3rem 0;">
<b style="color:#1a5488;">Clean&amp;QA</b> is an all-in-one tool for clinical, pharma and biomedical
localization. It <b>anonymizes</b> sensitive data, <b>cleans</b> your translation memories
and runs an independent <b>quality check</b> on bilingual memoQ files
(<b>MQXLIFF</b>), translation memories (<b>TMX</b>) and monolingual Word
documents (<b>.docx</b>). The original XML structure and
all inline tags are always preserved, so files reimport cleanly into memoQ.
</p>

<div style="background:#f4f8fb; border-left:4px solid #0e7bc0; border-radius:8px; padding:1rem 1.2rem; margin-bottom:1.4rem;">
  <div style="font-weight:700; color:#1a5488; font-size:1.05rem; margin-bottom:0.6rem;">🚀 Quick start</div>
  <ol style="margin:0; padding-left:1.2rem; line-height:1.85;">
    <li><b>Upload</b> one or more <b>.mqxliff</b> or <b>.tmx</b> files (you can mix both).</li>
    <li>In the sidebar, set your <b>replacement token</b> (default <code>███</code>) and turn the
        anonymization <b>layers</b> and <b>filters</b> on or off.</li>
    <li>Click <b>Process files</b> and watch the replacement statistics.</li>
    <li>Check the result in <b>Preview</b>, resolve <b>Duplicates</b> and, if you want,
        run the <b>QA Check</b>.</li>
    <li><b>Download</b> the anonymized files, a clean TMX and an Excel report.</li>
  </ol>
</div>

<div style="font-weight:700; color:#1a5488; font-size:1.05rem; margin-bottom:0.7rem;">🗂️ The five tabs</div>
<div style="display:grid; grid-template-columns:1fr 1fr; gap:0.55rem; margin-bottom:1.4rem;">
  <div style="border:1px solid #bcbdbe; border-radius:8px; padding:0.65rem 0.85rem;">
    <b style="color:#134277;">📤 Upload</b><br><span style="color:#5e5f6b;">Add your files and start processing. Optionally load a custom blacklist of terms to redact.</span></div>
  <div style="border:1px solid #bcbdbe; border-radius:8px; padding:0.65rem 0.85rem;">
    <b style="color:#134277;">📝 Preview</b><br><span style="color:#5e5f6b;">Compare every segment before / after, with anonymization stats. Mark junk segments to <i>Skip</i> so they stay untouched.</span></div>
  <div style="border:1px solid #bcbdbe; border-radius:8px; padding:0.65rem 0.85rem;">
    <b style="color:#134277;">🔄 Duplicates</b><br><span style="color:#5e5f6b;">Spot identical entries, translation conflicts and similar variants, and choose which ones to keep.</span></div>
  <div style="border:1px solid #bcbdbe; border-radius:8px; padding:0.65rem 0.85rem;">
    <b style="color:#134277;">🛡️ QA Check</b><br><span style="color:#5e5f6b;">Run 35 quality checks on a bilingual file, or a monolingual Word <b>.docx</b>, fully independent from the anonymization pipeline.</span></div>
  <div style="border:1px solid #bcbdbe; border-radius:8px; padding:0.65rem 0.85rem; grid-column:span 2;">
    <b style="color:#134277;">📥 Download</b><br><span style="color:#5e5f6b;">Get the cleaned files (single file or a ZIP for batches), a clean TMX and an Excel report of every excluded segment.</span></div>
</div>

<div style="font-weight:700; color:#1a5488; font-size:1.05rem; margin-bottom:0.5rem;">🛡️ Anonymization layers</div>
<p style="color:#5e5f6b; line-height:1.55; margin:0 0 0.7rem 0;">
  Six layers run in sequence; each protects what the previous one already redacted, so nothing is
  double-redacted. Turn on only the ones you need. A detected term is replaced <b>consistently</b>
  across the whole document and in <b>both languages</b>.
</p>
<table style="width:100%; border-collapse:collapse; margin-bottom:1.4rem; font-size:0.9rem;">
  <tr><td style="padding:0.35rem 0; width:40%; vertical-align:top;"><b>Safe Regex</b> <span style="background:#d7f0db; color:#1e7e34; border-radius:10px; padding:0 7px; font-size:0.72rem;">ON</span></td><td style="color:#5e5f6b;">Emails, phone numbers, URLs, IBAN, DNI/NIF, NCT and structured IDs, addresses and titled person names.</td></tr>
  <tr><td style="padding:0.35rem 0; vertical-align:top;"><b>Clinical ID Regex</b></td><td style="color:#5e5f6b;">Clinical-trial identifiers: NCT, EudraCT, Protocol, Subject and Site IDs.</td></tr>
  <tr><td style="padding:0.35rem 0; vertical-align:top;"><b>Custom dictionary</b> <span style="background:#d7f0db; color:#1e7e34; border-radius:10px; padding:0 7px; font-size:0.72rem;">ON</span></td><td style="color:#5e5f6b;">Your own blacklist of project-specific terms (brands, product names…) to always redact.</td></tr>
  <tr><td style="padding:0.35rem 0; vertical-align:top;"><b>Proper Names</b> <span style="background:#fbe3ee; color:#c2185b; border-radius:10px; padding:0 7px; font-size:0.72rem;">OFF</span></td><td style="color:#5e5f6b;">Rule-based person-name detection using multilingual first/last name lists (ES, EN, FR, DE, IT, PT, NL) plus hospital and institution patterns.</td></tr>
</table>

<div style="font-weight:700; color:#1a5488; font-size:1.05rem; margin-bottom:0.5rem;">🧹 Smart filters &amp; duplicates</div>
<p style="color:#5e5f6b; line-height:1.55; margin:0 0 0.5rem 0;">
  Filters keep low-value or over-redacted segments out of your translation memory:
</p>
<ul style="margin:0 0 0.8rem 0; padding-left:1.2rem; line-height:1.75; color:#130e45;">
  <li><b>Exclude short segments</b>: drops junk (only numbers/symbols or fewer than the chosen words). Override per segment with <i>Skip</i>.</li>
  <li><b>Exclude short anonymized segments</b>: removes segments left too short after redaction.</li>
  <li><b>Exclude heavily anonymized</b>: segments redacted above a chosen % are kept out of the TM (shown with a red dashed border and a % badge).</li>
</ul>
<p style="color:#5e5f6b; line-height:1.55; margin:0 0 1.4rem 0;">
  The <b>Duplicates</b> tab groups entries into <b>Identical</b> (same source &amp; target),
  <b>Same source, different target</b> (translation conflicts) and <b>Similar source</b>
  (fuzzy variants, with two adjustable sliders). Keep or discard each group before exporting a clean TMX.
</p>

<div style="background:#f7f3fb; border-left:4px solid #6c3483; border-radius:8px; padding:1rem 1.2rem; margin-bottom:1.4rem;">
  <div style="font-weight:700; color:#6c3483; font-size:1.05rem; margin-bottom:0.5rem;">✅ QA Check</div>
  <p style="margin:0 0 0.7rem 0; color:#130e45; line-height:1.6;">
    An independent quality check with <b>35 deterministic checks</b> across 5 groups
    (31 on by default, 4 opt-in). It flags real issues without false positives: empty or
    untranslated targets, number and date mismatches, broken or missing tags, forbidden /
    glossary terms and spelling. You can <b>edit any target inline</b>, then download the
    corrected file plus a CSV and an interactive <b>HTML report</b>. Spelling is opt-in:
    5 languages are bundled and ready to use (EN/ES/FR/DE/IT), and 33 more are
    downloaded automatically the first time their language is detected.
  </p>
  <p style="margin:0; color:#130e45; line-height:1.6;">
    It also accepts a standalone <b>Word (.docx)</b> document in <b>monolingual</b> mode:
    you pick the language and only the single-text checks run (spelling, repeated words,
    double spaces, brackets…). Output is the <b>HTML report</b> only — you fix the document yourself.
  </p>
  <div style="line-height:2;">
    <span style="background:#ffd9b3; color:#d35400; border-radius:12px; padding:2px 10px; font-size:0.8rem; font-weight:600;">Content</span>
    <span style="background:#cce0ff; color:#1f4e79; border-radius:12px; padding:2px 10px; font-size:0.8rem; font-weight:600;">Numeric Elements</span>
    <span style="background:#e8d4b8; color:#6b4423; border-radius:12px; padding:2px 10px; font-size:0.8rem; font-weight:600;">Tags</span>
    <span style="background:#e8d5f2; color:#6c3483; border-radius:12px; padding:2px 10px; font-size:0.8rem; font-weight:600;">Terminology</span>
    <span style="background:#f4cce0; color:#a01060; border-radius:12px; padding:2px 10px; font-size:0.8rem; font-weight:600;">Spelling</span>
  </div>
</div>

<div style="background:#fff8e6; border-left:4px solid #e0a800; border-radius:8px; padding:0.9rem 1.2rem; margin-bottom:0.8rem;">
  <b style="color:#8a6d00;">💡 Tip:</b>
  <span style="color:#5e5f6b;"> turn on <b>Safe Regex</b> + <b>Clinical ID Regex</b> +
  <b>Custom dictionary</b>, and load your project blacklist for brand and product names.</span>
</div>


</div>
        """,
        unsafe_allow_html=True,
    )


def main():
    st.markdown("""
    <style>
    .st-key-app_header {
        background: linear-gradient(135deg, #0e7bc0 0%, #1a5488 100%);
        padding: 1.75rem 2rem;
        border-radius: 12px;
        margin-bottom: 2rem;
        box-shadow: 0 4px 15px rgba(26, 84, 136, 0.3);
    }
    .st-key-app_header .stButton { display: flex; justify-content: flex-end; }
    .st-key-app_header .stButton button {
        background: rgba(255,255,255,0.18);
        color: #ffffff;
        border: 1px solid rgba(255,255,255,0.45);
        font-weight: 600;
        border-radius: 50%;
        width: 44px;
        height: 44px;
        min-height: 44px;
        padding: 0;
        line-height: 1;
    }
    .st-key-app_header .stButton button:hover {
        background: rgba(255,255,255,0.32);
        color: #ffffff;
        border-color: rgba(255,255,255,0.85);
    }
    </style>
    """, unsafe_allow_html=True)
    with st.container(key="app_header"):
        _title_col, _help_col = st.columns([5, 1], vertical_alignment="center")
        with _title_col:
            st.markdown("""
        <h1 style="color: white !important; margin: 0; font-size: 2.5rem; font-weight: 700; letter-spacing: 8px;">
            Clean&QA <span style="background: rgba(255,255,255,0.2); color: rgba(255,255,255,0.95); font-size: 0.9rem; font-weight: 500; padding: 3px 10px; border-radius: 20px; vertical-align: middle; letter-spacing: 1px; border: 1px solid rgba(255,255,255,0.3);">v6.3</span>
        </h1>
        <p style="color: rgba(255,255,255,0.95) !important; margin: 0.5rem 0 0 0; font-size: 1.1rem; font-weight: 400; letter-spacing: 1px;">
            Anonymize, clean &amp; run QA on your files
        </p>
            """, unsafe_allow_html=True)
        with _help_col:
            if st.button("❓", key="open_help_dialog", help="Help & quick guide"):
                _render_help_dialog()
    
    with st.sidebar:
        st.markdown("### ⚙️ Settings")
        replacement_token = st.text_input("Replacement token", value="███")
        st.session_state['replacement_token'] = replacement_token
        process_source = True
        process_target = True
        st.markdown('<hr class="sidebar-divider">', unsafe_allow_html=True)
        st.markdown("#### Multilingual layers")
        use_safe_regex = st.checkbox("Safe Regex", value=True, help="Emails, phones, URLs, IDs, addresses, titled names, etc.")
        use_dictionary = st.checkbox("Custom dictionary", value=True, help="Blacklist: manually loaded terms to anonymize")
        use_proper_names = st.checkbox("Proper Names", value=False, help="Rule-based person names (multilingual first/last name lists), hospitals and institutions.")
        st.markdown('<hr class="sidebar-divider">', unsafe_allow_html=True)
        st.markdown("#### EN > ES layers")
        use_regex = st.checkbox("Clinical ID Regex", value=False, help="NCT IDs, EudraCT, Protocol IDs, Subject IDs, etc.")
        st.markdown('<hr class="sidebar-divider">', unsafe_allow_html=True)
        st.markdown("#### Filters")
        dedup_tmx = st.checkbox("Deduplicate TMX segments", value=True, help="Find identical, same-source, and similar-source segments in the Duplicates tab")
        if dedup_tmx:
            variant_source_threshold = st.slider("Variant source similarity (≥%)", 50, 100, 85, 5, help="Only affects the 'Similar source' section. How alike two sources must be to count as variants. The Identical and Same-source sections always use exact matching.")
            variant_target_max = st.slider("Variant max. target similarity (≤%)", 50, 100, 100, 5, help="Secondary filter for the 'Similar source' section. Lower it to show only variants whose translations also differ. 100% = show all.")
        else:
            variant_source_threshold = 85
            variant_target_max = 100
        st.session_state['dedup_tmx'] = dedup_tmx
        st.session_state['variant_source_threshold'] = variant_source_threshold
        st.session_state['variant_target_max'] = variant_target_max
        filter_junk = st.checkbox("Exclude short segments", value=True, help="Removes short segments: less than min. words or only numbers/symbols")
        min_words_junk = st.slider("Minimum words (short segments)", 2, 10, 2, 1) if filter_junk else 2
        st.session_state['filter_junk'] = filter_junk
        st.session_state['min_words_junk'] = min_words_junk
        filter_short_segments = st.checkbox("Exclude short anon. segments", value=True, help="Excludes segments that are too short after anonymization")
        min_words = st.slider("Minimum words (anon. segments)", 2, 10, 5, 1) if filter_short_segments else 5
        st.session_state['filter_short_segments'] = filter_short_segments
        st.session_state['min_words'] = min_words
        exclude_modified_targets = st.checkbox("Exclude heavily anonymized", value=True)
        if exclude_modified_targets:
            exclusion_threshold = st.slider("Threshold (≥%)", 10, 90, 50, 5, help="Segments above this % anonymized are excluded")
            exclude_source_too = True
        else:
            exclusion_threshold = 20
            exclude_source_too = False
        st.session_state['exclusion_threshold'] = exclusion_threshold
        st.session_state['exclude_source_too'] = exclude_source_too
        st.session_state['exclude_modified_targets'] = exclude_modified_targets
    
    tab1, tab2, tab_dedup, tab_qa, tab3 = st.tabs(["📤 Upload", "📝 Preview", "🔄 Duplicates", "🛡️ QA Check", "📥 Download"])
    
    with tab1:
        st.markdown("### Upload MQXLIFF / TMX / Word files")
        
        mqxliff_files = st.file_uploader(
            "Select one or more .mqxliff, .tmx or .docx files",
            type=["mqxliff", "tmx", "docx"],
            accept_multiple_files=True,
            help="You can upload multiple files for batch processing. Word documents are anonymized in place (monolingual)."
        )
        
        docx_anonym_lang = "es"
        if mqxliff_files:
            st.success(f"✅ {len(mqxliff_files)} file(s) loaded")
            for f in mqxliff_files:
                ext = f.name.rsplit(".", 1)[-1].upper() if "." in f.name else "?"
                st.write(f"- {f.name} ({f.size / 1024:.1f} KB) — {ext}")
            
            has_docx_upload = any(f.name.lower().endswith(".docx") for f in mqxliff_files)
            if has_docx_upload:
                _docx_langs = {
                    "Español (es)": "es",
                    "English (en)": "en",
                    "Français (fr)": "fr",
                    "Deutsch (de)": "de",
                    "Italiano (it)": "it",
                    "Português (pt)": "pt",
                    "Català (ca)": "ca",
                    "Nederlands (nl)": "nl",
                    "Polski (pl)": "pl",
                }
                _docx_lang_label = st.selectbox(
                    "🌐 Word document language",
                    options=list(_docx_langs.keys()),
                    index=0,
                    key="docx_anonym_lang_label",
                    help="Word files carry no reliable language metadata; this feeds the language-aware anonymization layers."
                )
                docx_anonym_lang = _docx_langs[_docx_lang_label]
        
        st.markdown("---")
        st.markdown("### Custom dictionary (blacklist)")
        
        st.markdown("""
        <div class="info-box">
            <strong>ℹ️ Terms that will be anonymized:</strong><br>
            These terms will be forcefully anonymized even if not detected automatically.<br>
            One term per line or separated by commas:<br>
            <code>Example</code><br>
            <code>Example project, info@example.com</code><br>
            <code>Brand®, www.example.com</code>
        </div>
        """, unsafe_allow_html=True)
        
        dictionary_file = st.file_uploader(
            "Upload TXT file with sensitive terms (optional)",
            type=["txt"],
            help="Custom terms to anonymize, applied to all active layers"
        )
        
        dictionary_terms = set()
        if dictionary_file:
            dict_raw = dictionary_file.read()
            if dict_raw[:3] == b'\xef\xbb\xbf':
                dict_raw = dict_raw[3:]
            elif dict_raw[:2] in (b'\xff\xfe', b'\xfe\xff'):
                dict_raw = dict_raw.decode('utf-16').encode('utf-8')
            try:
                content = dict_raw.decode("utf-8")
            except UnicodeDecodeError:
                content = dict_raw.decode("latin-1")
            dictionary_terms = load_dictionary_terms(content)
            st.success(f"✅ {len(dictionary_terms)} unique terms loaded")
            
            with st.expander("View loaded terms"):
                for term in sorted(dictionary_terms):
                    st.write(f"- {term}")
        
        st.markdown("### Protected terms (whitelist)")
        
        st.markdown("""
        <div class="info-box">
            <strong>ℹ️ Terms that will NOT be anonymized:</strong><br>
            These terms will be preserved even if detected.<br>
            One term per line or separated by commas.
        </div>
        """, unsafe_allow_html=True)
        
        whitelist_file = st.file_uploader(
            "Upload TXT file with protected terms (optional)",
            type=["txt"],
            help="Terms that should never be anonymized",
            key="whitelist_uploader"
        )
        
        whitelist_terms = set()
        if whitelist_file:
            wl_raw = whitelist_file.read()
            if wl_raw[:3] == b'\xef\xbb\xbf':
                wl_raw = wl_raw[3:]
            elif wl_raw[:2] in (b'\xff\xfe', b'\xfe\xff'):
                wl_raw = wl_raw.decode('utf-16').encode('utf-8')
            try:
                wl_content = wl_raw.decode("utf-8")
            except UnicodeDecodeError:
                wl_content = wl_raw.decode("latin-1")
            whitelist_terms = load_dictionary_terms(wl_content)
            st.success(f"🛡️ {len(whitelist_terms)} protected terms loaded")
            
            with st.expander("View protected terms"):
                for term in sorted(whitelist_terms):
                    st.write(f"- {term}")
        
        if mqxliff_files:
            st.markdown("---")
            if st.button("🚀 Process files", type="primary", use_container_width=True):
                process_files(
                    mqxliff_files, replacement_token, process_source, process_target,
                    use_safe_regex, use_regex, use_proper_names,
                    use_dictionary, dictionary_terms, whitelist_terms,
                    docx_lang=docx_anonym_lang
                )
    
    with tab2:
        if "previews" in st.session_state and st.session_state.previews:
            st.markdown("### Changes preview")
            
            col_search, col_show_junk = st.columns([3, 1])
            with col_search:
                search_term = st.text_input(
                    "🔍 Search in preview",
                    placeholder="Type to filter changes...",
                    key="preview_search"
                )
            with col_show_junk:
                show_junk_in_preview = st.checkbox(
                    "Show short",
                    value=False,
                    key="show_junk_preview",
                    help="Show/hide short and short anon. segments in the preview (does not affect downloads)"
                )
            
            replacement_token = st.session_state.get('replacement_token', '███')
            threshold = st.session_state.get('exclusion_threshold', 20)
            exclude_enabled = st.session_state.get('exclude_modified_targets', False)
            exclude_source = st.session_state.get('exclude_source_too', False)
            
            if 'excluded_segments' not in st.session_state:
                st.session_state['excluded_segments'] = {}
            if 'no_anon_segments' not in st.session_state:
                st.session_state['no_anon_segments'] = {}
            if 'skip_junk_segments' not in st.session_state:
                st.session_state['skip_junk_segments'] = {}
            
            filter_junk = st.session_state.get('filter_junk', False)
            min_words_junk = st.session_state.get('min_words_junk', 2)
            filter_short = st.session_state.get('filter_short_segments', False)
            min_words = st.session_state.get('min_words', 5)
            
            total_matches = 0
            candidates_count = 0
            candidate_keys = set()
            junk_segments_count = 0
            short_segments_count = 0
            file_index = 0
            
            for filename, file_previews in st.session_state.previews.items():
                filtered_previews = []
                for preview in file_previews:
                    is_junk = is_junk_segment(preview, min_words_junk)
                    segment_key_junk = f"{filename}_{preview['segment']}"
                    
                    if is_junk:
                        is_skipped = st.session_state.get(f"skipjunk_{segment_key_junk}", False)
                        if not is_skipped:
                            junk_segments_count += 1
                        if show_junk_in_preview:
                            preview['_is_junk'] = True
                            preview['_junk_skipped'] = is_skipped
                            filtered_previews.append(preview)
                        continue
                    
                    preview['_is_junk'] = False
                    is_short_anon = filter_short and segment_word_count(preview) < min_words
                    if is_short_anon:
                        is_short_skipped = st.session_state.get(f"skipshort_{filename}_{preview['segment']}", False)
                        if not is_short_skipped:
                            short_segments_count += 1
                        preview['_is_short_anon'] = True
                        preview['_short_anon_skipped'] = is_short_skipped
                        if show_junk_in_preview:
                            filtered_previews.append(preview)
                        continue
                    
                    if not preview.get('changed', True):
                        continue
                    
                    has_token = (replacement_token in preview.get('source_after', '') or 
                                replacement_token in preview.get('target_after', ''))
                    if not has_token:
                        continue
                    
                    if search_term:
                        search_lower = search_term.lower()
                        if (search_lower in preview['source_before'].lower() or
                            search_lower in preview['source_after'].lower() or
                            search_lower in preview['target_before'].lower() or
                            search_lower in preview['target_after'].lower()):
                            filtered_previews.append(preview)
                    else:
                        filtered_previews.append(preview)
                
                non_junk_count = sum(1 for p in filtered_previews if not p.get('_is_junk', False) and not p.get('_is_short_anon', False))
                junk_in_file = sum(1 for p in filtered_previews if p.get('_is_junk', False))
                short_anon_in_file = sum(1 for p in filtered_previews if p.get('_is_short_anon', False))
                
                if filtered_previews:
                    if file_index > 0:
                        st.markdown('<hr style="border:none;border-top:2px solid #1a5488;margin:1.5rem 0;">', unsafe_allow_html=True)
                    file_index += 1
                    label_parts = []
                    if non_junk_count > 0:
                        label_parts.append(f"{non_junk_count} affected")
                    if junk_in_file > 0:
                        label_parts.append(f"{junk_in_file} short")
                    if short_anon_in_file > 0:
                        label_parts.append(f"{short_anon_in_file} short anon.")
                    total_matches += non_junk_count
                    with st.expander(f"📄 {filename} ({', '.join(label_parts)} segments)", expanded=True):
                        # Word (.docx) previews are informational only: the
                        # anonymized docx is downloaded as-is, so the Skip /
                        # no-anon / Exclude-TM toggles (XML-only post-passes)
                        # are hidden to avoid offering controls with no effect.
                        is_docx_preview = filename.lower().endswith(".docx")
                        for preview in filtered_previews:
                            preview_is_junk = preview.get('_is_junk', False)
                            segment_key = f"{filename}_{preview['segment']}"
                            
                            if preview_is_junk:
                                junk_skipped = preview.get('_junk_skipped', False)
                                col_header, col_skip_junk = st.columns([4, 1])
                                with col_header:
                                    if junk_skipped:
                                        st.markdown(
                                            f'<div style="background:#d4edda;border:1px solid #28a745;border-radius:6px;padding:0.5rem 0.8rem;margin:0.3rem 0;color:#155724;">'
                                            f'<strong>Segment {preview["segment"]}</strong> '
                                            f'<span style="background:#28a745;color:white;padding:2px 6px;border-radius:4px;font-size:0.75rem;">✓ Kept</span>'
                                            f'<br><small>Source: {preview["source_before"][:80] or "(empty)"} | Target: {preview["target_before"][:80] or "(empty)"}</small>'
                                            f'</div>',
                                            unsafe_allow_html=True
                                        )
                                    else:
                                        st.markdown(
                                            f'<div style="background:#e2e3e5;border:1px solid #6c757d;border-radius:6px;padding:0.5rem 0.8rem;margin:0.3rem 0;color:#495057;">'
                                            f'<strong>Segment {preview["segment"]}</strong> '
                                            f'<span style="background:#6c757d;color:white;padding:2px 6px;border-radius:4px;font-size:0.75rem;">🗑️ Short</span>'
                                            f'<br><small>Source: {preview["source_before"][:80] or "(empty)"} | Target: {preview["target_before"][:80] or "(empty)"}</small>'
                                            f'</div>',
                                            unsafe_allow_html=True
                                        )
                                with col_skip_junk:
                                    if not is_docx_preview:
                                        cb_key = f"skipjunk_{segment_key}"
                                        if cb_key not in st.session_state:
                                            st.session_state[cb_key] = False
                                        skip_junk = st.checkbox(
                                            "Skip",
                                            key=cb_key,
                                            help="Keep this segment (don't exclude as short)"
                                        )
                                        st.session_state['skip_junk_segments'][segment_key] = skip_junk
                                continue
                            
                            preview_is_short_anon = preview.get('_is_short_anon', False)
                            if preview_is_short_anon:
                                short_anon_skipped = preview.get('_short_anon_skipped', False)
                                col_header, col_skip_short = st.columns([4, 1])
                                with col_header:
                                    if short_anon_skipped:
                                        st.markdown(
                                            f'<div style="background:#d4edda;border:1px solid #28a745;border-radius:6px;padding:0.5rem 0.8rem;margin:0.3rem 0;color:#155724;">'
                                            f'<strong>Segment {preview["segment"]}</strong> '
                                            f'<span style="background:#28a745;color:white;padding:2px 6px;border-radius:4px;font-size:0.75rem;">✓ Kept</span>'
                                            f'<br><small>Source: {preview["source_after"][:80] or "(empty)"} | Target: {preview["target_after"][:80] or "(empty)"}</small>'
                                            f'</div>',
                                            unsafe_allow_html=True
                                        )
                                    else:
                                        st.markdown(
                                            f'<div style="background:#fff3cd;border:1px solid #ffc107;border-radius:6px;padding:0.5rem 0.8rem;margin:0.3rem 0;color:#856404;">'
                                            f'<strong>Segment {preview["segment"]}</strong> '
                                            f'<span style="background:#e0a800;color:white;padding:2px 6px;border-radius:4px;font-size:0.75rem;">✂️ Short anon.</span>'
                                            f'<br><small>Source: {preview["source_after"][:80] or "(empty)"} | Target: {preview["target_after"][:80] or "(empty)"}</small>'
                                            f'</div>',
                                            unsafe_allow_html=True
                                        )
                                with col_skip_short:
                                    if not is_docx_preview:
                                        cb_key = f"skipshort_{segment_key}"
                                        if cb_key not in st.session_state:
                                            st.session_state[cb_key] = False
                                        skip_short = st.checkbox(
                                            "Skip",
                                            key=cb_key,
                                            help="Keep this segment (don't exclude as short anon.)"
                                        )
                                        if 'skip_short_segments' not in st.session_state:
                                            st.session_state['skip_short_segments'] = {}
                                        st.session_state['skip_short_segments'][segment_key] = skip_short
                                continue
                            
                            src_before = preview.get('source_before', '')
                            tgt_before = preview.get('target_before', '')
                            src_after = preview.get('source_after', '')
                            tgt_after = preview.get('target_after', '')
                            if src_before:
                                source_pct = (1 - len(src_after.replace(replacement_token, '')) / max(len(src_before), 1)) * 100
                            else:
                                source_pct = 0
                            if tgt_before:
                                target_pct = (1 - len(tgt_after.replace(replacement_token, '')) / max(len(tgt_before), 1)) * 100
                            else:
                                target_pct = 0
                            max_pct = max(source_pct, target_pct)
                            # docx never enters the TM, so it can never be a
                            # TM-exclusion candidate (badge, checkbox, counter).
                            is_candidate = max_pct >= threshold and not is_docx_preview
                            
                            if is_candidate:
                                candidates_count += 1
                                candidate_keys.add(segment_key)
                            
                            has_changes = (preview['source_before'] != preview['source_after'] or 
                                         preview['target_before'] != preview['target_after'])
                            
                            col_header, col_no_anon, col_exclude = st.columns([3, 1, 1])
                            with col_header:
                                header_placeholder = st.empty()
                            
                            with col_no_anon:
                                if has_changes and not is_docx_preview:
                                    no_anon_val = st.session_state['no_anon_segments'].get(segment_key, False)
                                    no_anon = st.checkbox(
                                        "Skip",
                                        value=no_anon_val,
                                        key=f"noanon_{segment_key}",
                                        help="Keep original text without anonymizing"
                                    )
                                    st.session_state['no_anon_segments'][segment_key] = no_anon
                            
                            with col_exclude:
                                if is_candidate and exclude_enabled:
                                    default_val = st.session_state['excluded_segments'].get(segment_key, True)
                                    exclude_this = st.checkbox(
                                        "Exclude TM",
                                        value=default_val,
                                        key=f"excl_{segment_key}"
                                    )
                                    st.session_state['excluded_segments'][segment_key] = exclude_this
                            
                            is_no_anon = st.session_state['no_anon_segments'].get(segment_key, False)
                            
                            segment_header = f"**Segment {preview['segment']}**"
                            if is_no_anon:
                                segment_header += ' <span style="background:#28a745;color:white;padding:2px 6px;border-radius:4px;font-size:0.75rem;">✓ Not anonymized</span>'
                            elif is_candidate and exclude_enabled:
                                segment_header += f' <span class="exclude-badge">⚠️ {max_pct:.0f}% → exclusion</span>'
                            elif max_pct > 0:
                                segment_header += f' <span style="background:#c8ccd0;color:#333;padding:2px 6px;border-radius:4px;font-size:0.75rem;">{max_pct:.0f}% anonymized</span>'
                            header_placeholder.markdown(segment_header, unsafe_allow_html=True)
                            
                            if preview['source_before'] != preview['source_after']:
                                col1, col2 = st.columns(2)
                                with col1:
                                    st.markdown("**Source - Before:**")
                                    st.markdown(f'<div class="before-text">{preview["source_before"]}</div>', 
                                              unsafe_allow_html=True)
                                with col2:
                                    if is_no_anon:
                                        st.markdown("**Source - After** _(no changes)_:")
                                        st.markdown(f'<div class="before-text" style="border-color:#28a745;">{preview["source_before"]}</div>', 
                                                  unsafe_allow_html=True)
                                    else:
                                        label_src = "**Source - After:**"
                                        show_excluded = is_candidate and exclude_enabled and exclude_source and st.session_state['excluded_segments'].get(segment_key, True)
                                        if show_excluded:
                                            label_src = "**Source - After** _(will be excluded)_:"
                                        st.markdown(label_src)
                                        css_class = "after-text excluded-segment" if show_excluded else "after-text"
                                        st.markdown(f'<div class="{css_class}">{preview["source_after"]}</div>', 
                                                  unsafe_allow_html=True)
                            
                            if preview['target_before'] != preview['target_after']:
                                col1, col2 = st.columns(2)
                                with col1:
                                    st.markdown("**Target - Before:**")
                                    st.markdown(f'<div class="before-text">{preview["target_before"]}</div>', 
                                              unsafe_allow_html=True)
                                with col2:
                                    if is_no_anon:
                                        st.markdown("**Target - After** _(no changes)_:")
                                        st.markdown(f'<div class="before-text" style="border-color:#28a745;">{preview["target_before"]}</div>', 
                                                  unsafe_allow_html=True)
                                    else:
                                        label = "**Target - After:**"
                                        show_excluded = is_candidate and exclude_enabled and st.session_state['excluded_segments'].get(segment_key, True)
                                        if show_excluded:
                                            label = "**Target - After** _(will be excluded)_:"
                                        st.markdown(label)
                                        css_class = "after-text excluded-segment" if show_excluded else "after-text"
                                        st.markdown(f'<div class="{css_class}">{preview["target_after"]}</div>', 
                                                  unsafe_allow_html=True)
                            
                            st.markdown("---")
            
            if filter_junk and junk_segments_count > 0:
                st.info(f"ℹ️ {junk_segments_count} short segments excluded from TM (<{min_words_junk} words or only numbers/symbols)")
            
            if filter_short and short_segments_count > 0:
                st.info(f"ℹ️ {short_segments_count} short anonymized segments excluded from TM (less than {min_words} words)")
            
            if exclude_enabled and candidates_count > 0:
                excluded_count = sum(1 for k, v in st.session_state['excluded_segments'].items() if v and k in candidate_keys)
                st.info(f"ℹ️ {excluded_count} heavily anonymized segments excluded from TM (threshold: {threshold}%)")
            
            if total_matches == 0:
                if search_term:
                    st.warning(f"No changes found containing '{search_term}'")
                else:
                    st.info("No anonymized segments to preview. Check that the appropriate layers are enabled and reprocess the files.")
        else:
            st.info("Upload and process files to see the changes preview")
    
    with tab_dedup:
        if "results" in st.session_state and st.session_state.results:
            _all_previews = st.session_state.get('previews', {})
            previews = bilingual_previews(_all_previews)
            if _all_previews and not previews:
                st.info("Duplicate detection applies to MQXLIFF/TMX files only. Word documents are anonymized in place.")
            dedup_tmx = st.session_state.get('dedup_tmx', True)

            # Segments that already carry an x-document canonical ID were
            # written by a previous cleaning session, i.e. the user already
            # reviewed and kept them. Groups where EVERY member has an ID
            # are hidden by default so users stop re-confirming the same
            # duplicates on every upload. Mixed groups (a new segment
            # duplicating an already-reviewed one) still show up.
            _canon_cache = st.session_state.get('dedup_canonical_cache')
            _canon_sig = st.session_state.get('process_token', 0)
            if _canon_cache and _canon_cache.get('sig') == _canon_sig:
                dedup_canonical_map = _canon_cache['map']
            else:
                dedup_canonical_map = extract_existing_canonical_ids(
                    st.session_state.get('originals', {}))
                st.session_state['dedup_canonical_cache'] = {
                    'sig': _canon_sig, 'map': dedup_canonical_map,
                }

            def _group_already_reviewed(group):
                return all(
                    dedup_canonical_map.get((m['file'], m['segment']))
                    for m in group['members'])
            variant_source_threshold = st.session_state.get('variant_source_threshold', 85)
            variant_target_max = st.session_state.get('variant_target_max', 100)
            no_anon_segments = st.session_state.get('no_anon_segments', {})
            filter_junk = st.session_state.get('filter_junk', False)
            min_words_junk = st.session_state.get('min_words_junk', 2)
            filter_short = st.session_state.get('filter_short_segments', False)
            min_words = st.session_state.get('min_words', 5)

            if not dedup_tmx:
                st.info("Deduplication is disabled. Enable it in the sidebar to detect duplicates.")
                st.session_state['dedup_keep'] = {}
                st.session_state['dedup_group_keys'] = {}
                st.session_state['conflict_keep'] = {}
                st.session_state['conflict_group_keys'] = {}
                st.session_state['variant_keep'] = {}
                st.session_state['variant_group_keys'] = {}
            else:
                no_anon_active = tuple(sorted(k for k, v in no_anon_segments.items() if v))
                skipjunk_active = tuple(sorted(
                    k for k in st.session_state
                    if k.startswith('skipjunk_') and st.session_state[k]
                ))
                dedup_sig = (
                    st.session_state.get('process_token', 0),
                    st.session_state.get('replacement_token', '███'),
                    variant_source_threshold, variant_target_max,
                    filter_junk, min_words_junk, filter_short, min_words,
                    no_anon_active, skipjunk_active,
                )
                dedup_cache = st.session_state.get('dedup_cache')
                if dedup_cache and dedup_cache.get('sig') == dedup_sig:
                    identical_groups, conflict_groups, variant_groups = dedup_cache['groups']
                else:
                    identical_groups, conflict_groups, variant_groups = detect_dedup_sections(
                        previews, no_anon_segments, filter_junk, min_words_junk,
                        filter_short, min_words, variant_source_threshold, variant_target_max
                    )
                    st.session_state['dedup_cache'] = {
                        'sig': dedup_sig,
                        'groups': (identical_groups, conflict_groups, variant_groups),
                    }

                reviewed_counts = {
                    'identical': sum(1 for g in identical_groups if _group_already_reviewed(g)),
                    'conflict': sum(1 for g in conflict_groups if _group_already_reviewed(g)),
                    'variant': sum(1 for g in variant_groups if _group_already_reviewed(g)),
                }
                total_reviewed = sum(reviewed_counts.values())
                show_reviewed = False
                if total_reviewed:
                    st.info(
                        f"✅ {total_reviewed} duplicate group(s) hidden because every segment "
                        "already has a canonical ID from a previous cleaning session "
                        "(you already reviewed and kept them)."
                    )
                    show_reviewed = st.checkbox(
                        "Show already-reviewed groups anyway",
                        value=False,
                        key="dedup_show_reviewed",
                        help=(
                            "Groups where all segments carry an x-document canonical ID "
                            "were kept in a previous session. Enable this to review them again."
                        ),
                    )
                if not show_reviewed:
                    identical_groups = [g for g in identical_groups if not _group_already_reviewed(g)]
                    conflict_groups = [g for g in conflict_groups if not _group_already_reviewed(g)]
                    variant_groups = [g for g in variant_groups if not _group_already_reviewed(g)]

                if not identical_groups:
                    st.success("No identical segments detected (same source and same target).")
                    st.session_state['dedup_keep'] = {}
                    st.session_state['dedup_group_keys'] = {}
                else:
                    _render_dedup_section(
                        identical_groups,
                        title=f"🔁 {len(identical_groups)} identical matches (same source & target)",
                        intro="These rows are exact duplicates. By default the first is kept and the rest are discarded from the clean TMX.",
                        keep_state_key='dedup_keep',
                        group_keys_state_key='dedup_group_keys',
                        radio_prefix='dedup_radio',
                        expander_label=lambda i, m, s: f"Match {i + 1}: {len(m)} identical segments",
                        show_member_sim=False,
                    )

                st.markdown("---")

                if not conflict_groups:
                    st.success("No same-source conflicts detected (same source, different target).")
                    st.session_state['conflict_keep'] = {}
                    st.session_state['conflict_group_keys'] = {}
                else:
                    _render_dedup_section(
                        conflict_groups,
                        title=f"⚠️ {len(conflict_groups)} same-source conflicts (same source, different target)",
                        intro="Same source translated differently. Choose which translation to keep; the others are discarded from the clean TMX.",
                        keep_state_key='conflict_keep',
                        group_keys_state_key='conflict_group_keys',
                        radio_prefix='conflict_radio',
                        expander_label=lambda i, m, s: f"Match {i + 1}: {len(m)} translations of the same source",
                        show_member_sim=False,
                    )

                st.markdown("---")

                if not variant_groups:
                    st.success("No similar-source variants detected at the current thresholds.")
                    st.session_state['variant_keep'] = {}
                    st.session_state['variant_group_keys'] = {}
                else:
                    variant_intro = (
                        f"Sources that are similar but not identical (≥{variant_source_threshold}% source similarity"
                        + (f", target similarity ≤{variant_target_max}%" if variant_target_max < 100 else "")
                        + "). Review and choose which to keep."
                    )
                    _render_dedup_section(
                        variant_groups,
                        title=f"🔎 {len(variant_groups)} similar-source variants",
                        intro=variant_intro,
                        keep_state_key='variant_keep',
                        group_keys_state_key='variant_group_keys',
                        radio_prefix='variant_radio',
                        expander_label=lambda i, m, s: f"Match {i + 1}: {len(m)} variants — {s:.0f}% source similarity",
                        show_member_sim=True,
                    )
        else:
            st.info("Process the files to see duplicate analysis")

    with tab_qa:
        _render_qa_check_tab()

    with tab3:
        if "results" in st.session_state and st.session_state.results:
            st.markdown("### Anonymization statistics")
            
            total_stats = {"safe_regex": 0, "regex_ct": 0, "proper_names": 0, "dictionary": 0}
            for stats in st.session_state.all_stats.values():
                for key in total_stats:
                    total_stats[key] += stats.get(key, 0)
            
            col1, col2, col3, col4 = st.columns(4)
            render_stat_card("Safe Regex", total_stats["safe_regex"], col1, "stat-card-safe-regex")
            render_stat_card("Regex CT IDs", total_stats["regex_ct"], col2, "stat-card-regex-ct")
            render_stat_card("Proper Names", total_stats.get("proper_names", 0), col3, "stat-card-proper-names")
            render_stat_card("Dictionary", total_stats["dictionary"], col4, "stat-card-dictionary")
            
            st.markdown("---")
            st.markdown("### Download anonymized files")
            
            exclude_modified_targets = st.session_state.get('exclude_modified_targets', False)
            exclusion_threshold = st.session_state.get('exclusion_threshold', 20)
            exclude_source_too = st.session_state.get('exclude_source_too', False)
            excluded_segments = st.session_state.get('excluded_segments', {})
            
            def get_junk_segment_keys(filename: str, file_previews: list, min_words_junk: int = 2) -> set:
                """Gets the keys of junk segments that should be cleared."""
                junk_keys = set()
                for preview in file_previews:
                    if is_junk_segment(preview, min_words_junk):
                        seg_key = f"{filename}_{preview['segment']}"
                        if not st.session_state.get(f"skipjunk_{seg_key}", False):
                            junk_keys.add(seg_key)
                return junk_keys
            
            def get_short_segment_keys(filename: str, file_previews: list, min_words: int) -> set:
                """Gets the keys of short segments that should be restored."""
                short_keys = set()
                for preview in file_previews:
                    if segment_word_count(preview) < min_words:
                        short_keys.add(f"{filename}_{preview['segment']}")
                return short_keys
            
            def apply_no_anon_segments(anon_content: bytes, orig_content: bytes, filename: str, no_anon_segs: dict) -> bytes:
                if not no_anon_segs:
                    return anon_content
                
                from lxml import etree
                is_tmx = filename.lower().endswith(".tmx")
                
                try:
                    anon_tree = etree.fromstring(anon_content)
                    orig_tree = etree.fromstring(orig_content)
                    
                    if is_tmx:
                        anon_units = anon_tree.xpath('//tu')
                        orig_units = orig_tree.xpath('//tu')
                    else:
                        nsmap = anon_tree.nsmap
                        default_ns = nsmap.get(None, '')
                        
                        if default_ns:
                            ns = {'x': default_ns}
                            anon_units = anon_tree.xpath('//x:trans-unit', namespaces=ns)
                            orig_units = orig_tree.xpath('//x:trans-unit', namespaces=ns)
                        else:
                            anon_units = anon_tree.xpath('//trans-unit')
                            orig_units = orig_tree.xpath('//trans-unit')
                    
                    replacement_token = st.session_state.get('replacement_token', '███')
                    
                    for idx, (anon_tu, orig_tu) in enumerate(zip(anon_units, orig_units)):
                        segment_num = idx + 1
                        segment_key = f"{filename}_{segment_num}"
                        
                        should_restore_only = no_anon_segs.get(segment_key, False)
                        
                        if should_restore_only:
                            if is_tmx:
                                anon_tuvs = anon_tu.xpath('tuv')
                                orig_tuvs = orig_tu.xpath('tuv')
                                for anon_tuv, orig_tuv in zip(anon_tuvs, orig_tuvs):
                                    anon_segs = anon_tuv.xpath('seg')
                                    orig_segs = orig_tuv.xpath('seg')
                                    for anon_seg, orig_seg in zip(anon_segs, orig_segs):
                                        anon_seg.getparent().replace(anon_seg, orig_seg)
                            else:
                                if default_ns:
                                    anon_sources = anon_tu.xpath('.//x:source', namespaces=ns)
                                    orig_sources = orig_tu.xpath('.//x:source', namespaces=ns)
                                    anon_targets = anon_tu.xpath('.//x:target', namespaces=ns)
                                    orig_targets = orig_tu.xpath('.//x:target', namespaces=ns)
                                else:
                                    anon_sources = anon_tu.xpath('.//source')
                                    orig_sources = orig_tu.xpath('.//source')
                                    anon_targets = anon_tu.xpath('.//target')
                                    orig_targets = orig_tu.xpath('.//target')
                                
                                for anon_src, orig_src in zip(anon_sources, orig_sources):
                                    anon_src.getparent().replace(anon_src, orig_src)
                                
                                for anon_tgt, orig_tgt in zip(anon_targets, orig_targets):
                                    anon_tgt.getparent().replace(anon_tgt, orig_tgt)
                    
                    return etree.tostring(anon_tree, encoding='utf-8', xml_declaration=True)
                except Exception:
                    # Malformed/encrypted XML — fall back to the anonymized
                    # content as-is so the user still gets a usable file.
                    return anon_content
            
            def _clear_element(elem):
                for child in list(elem):
                    elem.remove(child)
                elem.text = None
                elem.tail = None

            def prepare_download_content(content: bytes, filename: str, exclude_targets: bool, threshold: float, exclude_source: bool, excluded_segs: dict, short_segs: set = None, junk_segs: set = None) -> bytes:
                if short_segs is None:
                    short_segs = set()
                if junk_segs is None:
                    junk_segs = set()
                
                if not exclude_targets and not short_segs and not junk_segs:
                    return content
                
                from lxml import etree
                is_tmx = filename.lower().endswith(".tmx")
                
                try:
                    tree = etree.fromstring(content)
                    
                    if is_tmx:
                        trans_units = tree.xpath('//tu')
                    else:
                        nsmap = tree.nsmap
                        default_ns = nsmap.get(None, '')
                        if default_ns:
                            ns = {'x': default_ns}
                            trans_units = tree.xpath('//x:trans-unit', namespaces=ns)
                        else:
                            trans_units = tree.xpath('//trans-unit')
                    
                    replacement_token = st.session_state.get('replacement_token', '███')
                    
                    for idx, tu in enumerate(trans_units):
                        segment_num = idx + 1
                        segment_key = f"{filename}_{segment_num}"
                        
                        is_short = segment_key in short_segs
                        is_junk = segment_key in junk_segs
                        
                        if is_short or is_junk:
                            if is_tmx:
                                for tuv in tu.xpath('tuv'):
                                    for seg in tuv.xpath('seg'):
                                        _clear_element(seg)
                            else:
                                if default_ns:
                                    targets = tu.xpath('.//x:target', namespaces=ns)
                                    sources = tu.xpath('.//x:source', namespaces=ns)
                                else:
                                    targets = tu.xpath('.//target')
                                    sources = tu.xpath('.//source')
                                for elem in sources + targets:
                                    _clear_element(elem)
                            continue
                        
                        if not exclude_targets:
                            continue
                        
                        should_exclude_this = excluded_segs.get(segment_key, None)
                        if should_exclude_this is False:
                            continue
                        
                        if is_tmx:
                            target_tuvs = []
                            source_tuvs = []
                            for tuv in tu.xpath('tuv'):
                                tuv_lang = tuv.get("{http://www.w3.org/XML/1998/namespace}lang", tuv.get("lang", ""))
                                if tuv_lang.lower().startswith("es"):
                                    target_tuvs.extend(tuv.xpath('seg'))
                                elif tuv_lang.lower().startswith("en") and exclude_source:
                                    source_tuvs.extend(tuv.xpath('seg'))
                            
                            for seg in target_tuvs:
                                text_content = ''.join(seg.itertext())
                                words = text_content.split()
                                if replacement_token in text_content and words:
                                    redacted_pct = sum(1 for w in words if replacement_token in w) / len(words) * 100
                                    if redacted_pct >= threshold:
                                        _clear_element(seg)
                                        if exclude_source:
                                            for src_seg in source_tuvs:
                                                _clear_element(src_seg)
                        else:
                            if default_ns:
                                targets = tu.xpath('.//x:target', namespaces=ns)
                                sources = tu.xpath('.//x:source', namespaces=ns) if exclude_source else []
                            else:
                                targets = tu.xpath('.//target')
                                sources = tu.xpath('.//source') if exclude_source else []
                            
                            for target in targets:
                                text_content = ''.join(target.itertext())
                                words = text_content.split()
                                if replacement_token in text_content and words:
                                    redacted_pct = sum(1 for w in words if replacement_token in w) / len(words) * 100
                                    
                                    if redacted_pct >= threshold:
                                        _clear_element(target)
                                        
                                        if exclude_source:
                                            for source in sources:
                                                _clear_element(source)
                    
                    return etree.tostring(tree, encoding='utf-8', xml_declaration=True)
                except Exception:
                    # Malformed/encrypted XML — return the original bytes
                    # unmodified rather than crashing the download.
                    return content
            
            no_anon_segments = st.session_state.get('no_anon_segments', {})
            originals = st.session_state.get('originals', {})
            
            no_anon_count = sum(1 for v in no_anon_segments.values() if v)
            if no_anon_count > 0:
                st.markdown(f"""
                <div style="background-color: #d4edda; border: 1px solid #28a745; border-radius: 8px; padding: 1rem; margin: 0.5rem 0;">
                    <strong>✓ {no_anon_count} segments will keep their original text</strong><br>
                    <small>Marked as "Skip" in the Preview tab</small>
                </div>
                """, unsafe_allow_html=True)
            
            filter_junk = st.session_state.get('filter_junk', False)
            min_words_junk = st.session_state.get('min_words_junk', 2)
            filter_short = st.session_state.get('filter_short_segments', False)
            min_words = st.session_state.get('min_words', 5)
            previews = st.session_state.get('previews', {})
            # Word (.docx) files are monolingual and never enter the TM
            # pipeline (clean TMX, dedup, quality funnel).
            bilingual_prevs = bilingual_previews(previews)
            has_bilingual = bool(bilingual_prevs)
            
            dedup_tmx = st.session_state.get('dedup_tmx', True)
            
            existing_canonical_map = extract_existing_canonical_ids(originals)
            existing_canonical_count = len(existing_canonical_map)
            total_preview_segments = sum(len(fp) for fp in bilingual_prevs.values())
            all_have_canonical = existing_canonical_count >= total_preview_segments and existing_canonical_count > 0
            has_some_canonical = existing_canonical_count > 0
            segments_without_canonical = total_preview_segments - existing_canonical_count
            
            st.markdown("---")
            
            canonical_id_value = None
            if all_have_canonical:
                existing_ids = sorted(set(existing_canonical_map.values()))
                ids_display = ", ".join(existing_ids)
                st.markdown(f"""
                <div style="background-color: #d4edda; border: 1px solid #28a745; border-radius: 8px; padding: 1rem; margin: 0.5rem 0;">
                    <strong>🔒 All {existing_canonical_count} segments already have canonical IDs</strong><br>
                    <small>Existing IDs: {ids_display}</small>
                </div>
                """, unsafe_allow_html=True)
            else:
                if has_some_canonical:
                    existing_ids = sorted(set(existing_canonical_map.values()))
                    ids_display = ", ".join(existing_ids)
                    st.markdown(f"""
                    <div style="background-color: #fff3cd; border: 1px solid #ffc107; border-radius: 8px; padding: 1rem; margin: 0.5rem 0;">
                        <strong>⚠️ {existing_canonical_count} segments already have canonical IDs ({ids_display})</strong><br>
                        <small>{segments_without_canonical} segments without canonical ID</small>
                    </div>
                    """, unsafe_allow_html=True)
                
                if 'use_canonical' not in st.session_state:
                    st.session_state['use_canonical'] = True
                use_canonical = st.checkbox("Assign canonical ID to TMX segments", key='use_canonical',
                                            help="Each segment will carry a permanent document identifier for traceability")
                
                if use_canonical:
                    if 'canonical_counter' not in st.session_state:
                        st.session_state['canonical_counter'] = 1
                    suggested_id = f"TMX-{date.today().isoformat()}-{st.session_state['canonical_counter']:03d}"
                    canonical_id_value = st.text_input("Canonical ID", value=suggested_id, key='canonical_id_input',
                                                        help="All segments in this TMX will be tagged with this ID")
                    if canonical_id_value:
                        canonical_id_value = canonical_id_value.strip()
                    if not canonical_id_value:
                        st.warning("Please enter a canonical ID or uncheck the option.")
                        canonical_id_value = None
                    elif has_some_canonical:
                        st.info(f"Only {segments_without_canonical} new segments will receive this ID. Existing canonical IDs are preserved.")
            
            tmx_output_name = f"Redacted_TMX_{date.today().isoformat()}.tmx"
            
            clean_tmx_data, valid_segments, no_anon_in_tmx, excluded_ids, dedup_count, dedup_details, exclusion_breakdown = generate_clean_tmx(
                previews=bilingual_prevs,
                results=st.session_state.results,
                originals=originals,
                filter_junk=filter_junk,
                min_words_junk=min_words_junk,
                filter_short=filter_short,
                min_words=min_words,
                exclude_modified=exclude_modified_targets,
                exclusion_threshold=exclusion_threshold,
                excluded_segments=excluded_segments,
                no_anon_segments=no_anon_segments,
                dedup_tmx=dedup_tmx,
                dedup_keep_choices=st.session_state.get('dedup_keep', {}),
                tmx_filename=tmx_output_name,
                canonical_id=canonical_id_value,
                existing_canonical_map=existing_canonical_map
            )
            
            st.session_state['dedup_details'] = dedup_details
            st.session_state['exclusion_breakdown'] = exclusion_breakdown
            
            total_segs = sum(len(fp) for fp in bilingual_prevs.values())
            empty_total = total_segs - valid_segments - no_anon_in_tmx
            if empty_total > 0:
                breakdown_parts = []
                if len(exclusion_breakdown.get("short", [])) > 0:
                    breakdown_parts.append(f"{len(exclusion_breakdown['short'])} short")
                if len(exclusion_breakdown.get("short_anon", [])) > 0:
                    breakdown_parts.append(f"{len(exclusion_breakdown['short_anon'])} short anon.")
                if len(exclusion_breakdown.get("empty", [])) > 0:
                    breakdown_parts.append(f"{len(exclusion_breakdown['empty'])} empty")
                if len(exclusion_breakdown.get("heavy", [])) > 0:
                    breakdown_parts.append(f"{len(exclusion_breakdown['heavy'])} heavily anon.")
                if len(exclusion_breakdown.get("dedup", [])) > 0:
                    breakdown_parts.append(f"{len(exclusion_breakdown['dedup'])} duplicates")
                breakdown_text = " | ".join(breakdown_parts) if breakdown_parts else "Excluded by filters"
                st.markdown(f"""
                <div style="background-color: #e2e3e5; border: 1px solid #6c757d; border-radius: 8px; padding: 1rem; margin: 0.5rem 0;">
                    <strong>🗑️ {empty_total} segments excluded from clean TMX</strong> <small>(of {total_segs} total)</small><br>
                    <small>{breakdown_text}</small>
                </div>
                """, unsafe_allow_html=True)
            
            # Task #76 — TM Quality Summary (session-only, mirrors the
            # "TM Quality Report" sheet inside the changes report Excel).
            # Gated on bilingual files: Word documents never enter the TM
            # funnel, so a docx-only session has no TM quality to report.
            quality_metrics = None
            if has_bilingual:
                quality_metrics = collect_quality_metrics(
                    previews=bilingual_prevs,
                    all_stats={
                        fn: s for fn, s in st.session_state.get('all_stats', {}).items()
                        if not fn.lower().endswith(".docx")
                    },
                    exclusion_breakdown=exclusion_breakdown,
                    qa_results=st.session_state.get('qa_results')
                )
                qb = quality_metrics["batch"]
                with st.expander("📊 TM Quality Summary (this session)", expanded=False):
                    st.markdown(f"""
                    <div style="background-color: #d4edda; border: 1px solid #28a745; border-radius: 8px; padding: 1rem; margin: 0.5rem 0;">
                        <strong style="color:#1a5488; font-size:1.1rem;">Retention rate: {qb['retention_pct']:.1f}%</strong><br>
                        <small>{qb['final']} of {qb['raw']} raw segments entered the final clean TM</small>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    funnel_lines = [
                        ("Raw segments", qb['raw']),
                        ("Junk / short segments", qb['short']),
                        ("Short after anonymization", qb['short_anon']),
                        ("Empty after cleaning", qb['empty']),
                        ("Heavily anonymized", qb['heavy']),
                        ("Kept without anonymization", qb['no_anon']),
                        ("Duplicates discarded", qb['dedup_total']),
                        ("Final segments in clean TM", qb['final']),
                    ]
                    rows_html = "".join(
                        f"<tr><td style='padding:2px 12px 2px 0; color:#130e45;'>{label}</td>"
                        f"<td style='text-align:right; font-weight:{'bold' if label.startswith(('Raw', 'Final')) else 'normal'};'>{count}</td></tr>"
                        for label, count in funnel_lines
                    )
                    dd = (f"Duplicates: {qb['dedup_identical']} identical · "
                          f"{qb['dedup_conflict']} conflicts · {qb['dedup_variant']} variants")
                    qa_m = quality_metrics["qa"]
                    if qa_m["run"]:
                        qa_fname = html.escape(str(qa_m['filename'] or ''))
                        qa_line = (f"QA ({qa_fname}): {qa_m['total']} issues — " +
                                   " · ".join(f"{g} {qa_m['by_group'].get(g, 0)}" for g in QA_GROUP_ORDER))
                    else:
                        qa_line = "QA not run in this session"
                    st.markdown(f"""
                    <div style="background-color: #ffffff; border: 1px solid #bcbdbe; border-radius: 8px; padding: 1rem; margin: 0.5rem 0;">
                        <table style="width:100%; font-size:0.9rem;">{rows_html}</table>
                        <hr style="margin:0.5rem 0; border-color:#e0e3e4;">
                        <small>{dd}</small><br>
                        <small>Anonymization: {qb['anon_total']} replacements · {qb['unique_terms']} unique terms</small><br>
                        <small>{qa_line}</small>
                    </div>
                    """, unsafe_allow_html=True)
                    st.caption("The full report (per file, layers, speed and configuration) is included as the first sheet of the 📊 changes report Excel.")
            
            col_mqxliff, col_tmx_clean, col_excel = st.columns(3)
            
            with col_mqxliff:
                if len(st.session_state.results) == 1:
                    filename, content = list(st.session_state.results.items())[0]
                    orig_content = originals.get(filename, content)
                    if filename.lower().endswith(".docx"):
                        # Word files are already final: the XML-based
                        # no-anon/exclusion post-passes do not apply.
                        download_content = content
                        mime_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
                    else:
                        junk_segs = get_junk_segment_keys(filename, previews.get(filename, []), min_words_junk) if filter_junk else set()
                        short_segs = get_short_segment_keys(filename, previews.get(filename, []), min_words) if filter_short else set()
                        content = apply_no_anon_segments(content, orig_content, filename, no_anon_segments)
                        download_content = prepare_download_content(content, filename, exclude_modified_targets, exclusion_threshold, exclude_source_too, excluded_segments, short_segs, junk_segs)
                        mime_type = "application/xml"
                    file_ext = filename.rsplit(".", 1)[-1] if "." in filename else "xml"
                    redacted_filename = f"Redacted_{date.today().isoformat()}.{file_ext}"
                    st.download_button(
                        label=f"📥 Download {file_ext.upper()}",
                        data=download_content,
                        file_name=redacted_filename,
                        mime=mime_type,
                        use_container_width=True
                    )
                else:
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                        for idx, (filename, content) in enumerate(st.session_state.results.items()):
                            orig_content = originals.get(filename, content)
                            if filename.lower().endswith(".docx"):
                                download_content = content
                            else:
                                junk_segs = get_junk_segment_keys(filename, previews.get(filename, []), min_words_junk) if filter_junk else set()
                                short_segs = get_short_segment_keys(filename, previews.get(filename, []), min_words) if filter_short else set()
                                content = apply_no_anon_segments(content, orig_content, filename, no_anon_segments)
                                download_content = prepare_download_content(content, filename, exclude_modified_targets, exclusion_threshold, exclude_source_too, excluded_segments, short_segs, junk_segs)
                            file_ext = filename.rsplit(".", 1)[-1] if "." in filename else "xml"
                            zip_entry_name = f"Redacted_{idx + 1}_{date.today().isoformat()}.{file_ext}"
                            zf.writestr(zip_entry_name, download_content)
                    
                    st.download_button(
                        label=f"📥 Download ZIP ({len(st.session_state.results)} files)",
                        data=zip_buffer.getvalue(),
                        file_name=f"Redacted_{date.today().isoformat()}.zip",
                        mime="application/zip",
                        use_container_width=True
                    )
            
            with col_tmx_clean:
                if has_bilingual:
                    tmx_label = "📥 Download clean TMX"
                    tmx_help = "TMX without empty, excluded or filtered segments"
                    if st.download_button(
                        label=tmx_label,
                        data=clean_tmx_data,
                        file_name=tmx_output_name,
                        mime="application/xml",
                        use_container_width=True,
                        help=tmx_help
                    ):
                        if canonical_id_value and 'canonical_counter' in st.session_state:
                            st.session_state['canonical_counter'] += 1
                else:
                    st.info("Clean TMX applies to MQXLIFF/TMX files only.")
            
            with col_excel:
                file_canonical_map = {}
                for fname in st.session_state.get('originals', {}).keys():
                    cids = set()
                    file_existing_segments = set()
                    if existing_canonical_map:
                        for (f, seg), cid in existing_canonical_map.items():
                            if f == fname:
                                cids.add(cid)
                                file_existing_segments.add(seg)
                    if canonical_id_value:
                        file_previews = previews.get(fname, [])
                        has_new_segments = any(
                            p['segment'] not in file_existing_segments
                            for p in file_previews
                        )
                        if has_new_segments:
                            cids.add(canonical_id_value)
                    file_canonical_map[fname] = ", ".join(sorted(cids)) if cids else "None"
                
                excel_data = generate_changes_excel(
                    dedup_details=st.session_state.get('dedup_details', []),
                    exclusion_breakdown=st.session_state.get('exclusion_breakdown', None),
                    file_canonical_map=file_canonical_map if file_canonical_map else None,
                    quality_metrics=quality_metrics
                )
                st.download_button(
                    label="📊 Download changes report",
                    data=excel_data,
                    file_name=f"Report_{date.today().isoformat()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
        else:
            st.info("Process the files to download them")


def process_files(files, replacement_token, process_source, process_target,
                  use_safe_regex, use_regex, use_proper_names,
                  use_dictionary, dictionary_terms, whitelist_terms=None, docx_lang="es"):
    
    anonymizer = MQXLIFFAnonymizer(replacement_token=replacement_token)
    
    results = {}
    originals = {}
    all_stats = {}
    previews = {}

    seen_names = set()

    def _unique_name(name):
        """Return a collision-free display/key name. When several uploaded
        files share a basename, later ones get a ' (2)', ' (3)', ... suffix
        (before the extension) so they no longer overwrite each other in the
        results/originals/previews maps."""
        if name not in seen_names:
            seen_names.add(name)
            return name
        if "." in name:
            base, ext = name.rsplit(".", 1)
            fmt = lambda n: f"{base} ({n}).{ext}"
        else:
            fmt = lambda n: f"{name} ({n})"
        n = 2
        while fmt(n) in seen_names:
            n += 1
        new_name = fmt(n)
        seen_names.add(new_name)
        return new_name

    progress_bar = st.progress(0, text="Processing files...")
    
    for i, file in enumerate(files):
        unique_name = _unique_name(file.name)
        file_label = unique_name if len(unique_name) <= 40 else unique_name[:37] + "..."
        
        def make_progress_cb(file_idx, total_files, fname):
            last_pct = [-1]
            def cb(current, total):
                pct = int((current / total) * 100) if total > 0 else 0
                if pct == last_pct[0] and pct < 100:
                    return
                last_pct[0] = pct
                file_base = file_idx / total_files
                file_share = 1.0 / total_files
                segment_pct = current / total if total > 0 else 0
                overall = file_base + file_share * segment_pct
                progress_bar.progress(min(overall, 1.0), text=f"Processing {fname}... {pct}% ({current}/{total} segments)")
                time.sleep(0.01)
            return cb
        
        progress_cb = make_progress_cb(i, len(files), file_label)
        progress_cb(0, 1)
        
        try:
            _t0 = time.perf_counter()
            content = file.read()
            originals[unique_name] = content
            
            is_tmx = unique_name.lower().endswith(".tmx")
            is_docx = unique_name.lower().endswith(".docx")
            
            if is_docx:
                result_xml, stats, file_previews = anonymizer.anonymize_docx(
                    docx_bytes=content,
                    lang=docx_lang,
                    use_safe_regex=use_safe_regex,
                    use_regex=use_regex,
                    use_proper_names=use_proper_names,
                    use_dictionary=use_dictionary,
                    dictionary_terms=dictionary_terms,
                    whitelist_terms=whitelist_terms,
                    progress_callback=progress_cb
                )
            elif is_tmx:
                result_xml, stats, file_previews = anonymizer.anonymize_tmx(
                    xml_content=content,
                    process_source=process_source,
                    process_target=process_target,
                    use_safe_regex=use_safe_regex,
                    use_regex=use_regex,
                    use_proper_names=use_proper_names,
                    use_dictionary=use_dictionary,
                    dictionary_terms=dictionary_terms,
                    whitelist_terms=whitelist_terms,
                    progress_callback=progress_cb
                )
            else:
                result_xml, stats, file_previews = anonymizer.anonymize_mqxliff(
                    xml_content=content,
                    process_source=process_source,
                    process_target=process_target,
                    use_safe_regex=use_safe_regex,
                    use_regex=use_regex,
                    use_proper_names=use_proper_names,
                    use_dictionary=use_dictionary,
                    dictionary_terms=dictionary_terms,
                    whitelist_terms=whitelist_terms,
                    progress_callback=progress_cb
                )
            
            # Task #76 — session quality metrics: wall-clock per file,
            # per-file unique anonymized terms (terms_cache is reset per
            # file by reset_stats) and raw segment count.
            stats["wall_ms"] = round((time.perf_counter() - _t0) * 1000.0, 1)
            stats["unique_terms"] = len(anonymizer.terms_cache)
            stats["segments_total"] = len(file_previews)
            
            results[unique_name] = result_xml
            all_stats[unique_name] = stats
            previews[unique_name] = file_previews
            
        except Exception as e:
            st.error(f"Error processing {unique_name}: {str(e)}")
    
    progress_bar.empty()
    
    st.session_state.results = results
    st.session_state.originals = originals
    st.session_state.all_stats = all_stats
    st.session_state.previews = previews
    # Task #76 — snapshot of the anonymization layers used this session,
    # for the TM Quality Report configuration section.
    st.session_state['session_layers'] = {
        "Safe Regex": bool(use_safe_regex),
        "Clinical ID Regex": bool(use_regex),
        "Proper Names": bool(use_proper_names),
        "Custom dictionary": bool(use_dictionary),
    }
    st.session_state['use_canonical'] = True
    st.session_state['dedup_keep'] = {}
    st.session_state['dedup_group_keys'] = {}
    st.session_state['conflict_keep'] = {}
    st.session_state['conflict_group_keys'] = {}
    st.session_state['variant_keep'] = {}
    st.session_state['variant_group_keys'] = {}
    st.session_state['process_token'] = st.session_state.get('process_token', 0) + 1
    st.session_state['dedup_cache'] = None
    
    st.success(f"✅ Successfully processed {len(results)} file(s)")
    st.info("Go to the **Preview** and **Download** tabs to see the results")


if __name__ == "__main__":
    main()
